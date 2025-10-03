#!/usr/bin/env python3
"""
Project Financials Consultant Metrics Analyzer

Comprehensive analysis tool for consultant and solution architect performance
from Project Financials Excel data. Features include:

- Consultant efficiency scoring with composite ranking
- Solution architect volume-based performance analysis
- Advanced analytics: predictive modeling, anomaly detection, risk assessment
- Smart data processing: duplicate handling, comma-separated fields
- Detailed Excel exports with color-coded variance highlighting
- Configurable filtering and thresholds

Author: Ron Loge
Version: 2.0
Last Updated: 2025-07-03
"""

# Standard library imports
import pandas as pd  # Data manipulation and analysis
import numpy as np   # Numerical computing
import os           # Operating system interface
import sys          # System-specific parameters
import json         # JSON encoder/decoder
from pathlib import Path  # Object-oriented filesystem paths
import warnings
warnings.filterwarnings('ignore')  # Suppress pandas warnings for cleaner output

class ConsultantMetricsAnalyzer:
    """
    Main analyzer class that processes project financial data and generates
    comprehensive performance metrics for consultants and solution architects.
    
    The analyzer supports:
    - Multi-threshold scoring (efficiency vs success rates)
    - Composite scoring with volume bonuses
    - Advanced analytics with statistical analysis
    - Configurable filtering and data processing
    """
    def __init__(self, config_file='config.json'):
        """
        Initialize the analyzer with configuration settings.
        
        Args:
            config_file (str): Path to JSON configuration file
            
        Sets up:
        - Configuration loading from JSON
        - File paths for input data
        - Empty containers for processed data
        - Debug mode flag
        """
        # Load configuration from JSON file
        self.config = self.load_config(config_file)
        
        # Validate configuration
        if not self.validate_config():
            print("[ERROR] Configuration validation failed. Please fix the issues above.")
            sys.exit(1)
        
        # Extract file paths from configuration
        self.excel_file = self.config['files']['excel_file']
        self.engineers_file = self.config['files']['engineers_file']
        self.exclude_file = self.config['files']['exclude_file']
        self.solution_architects_file = self.config['files']['solution_architects_file']
        
        # Initialize data containers
        self.data = None  # Main DataFrame with project data (filtered)
        self.original_data = None  # Unfiltered data for customer analysis
        self.engineers_list = []  # List of tracked consultants
        self.exclusions = []  # List of (consultant, project) exclusions
        self.solution_architects_list = []  # List of tracked SAs
        
        # Results containers
        self.consultant_metrics = {}  # Consultant performance metrics
        self.sa_metrics = {}  # Solution architect performance metrics
        
        # Control flags
        self.show_composite = False  # Debug mode flag for showing composite scores
    
    def load_config(self, config_file):
        """
        Load configuration from JSON file.
        
        Args:
            config_file (str): Path to configuration file
            
        Returns:
            dict: Configuration dictionary
            
        Exits:
            System exit if config file not found
        """
        try:
            with open(config_file, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"[ERROR] Config file {config_file} not found")
            sys.exit(1)
    
    def validate_config(self):
        """
        Validate configuration settings and provide helpful error messages.
        
        Returns:
            bool: True if configuration is valid, False otherwise
        """
        valid = True
        
        # Validate thresholds
        if not self.validate_thresholds():
            valid = False
        
        # Validate data source
        if not self.validate_data_source():
            valid = False
        
        # Validate date filtering
        if not self.validate_date_filtering():
            valid = False
        
        # Validate DAS+ settings
        if not self.validate_das_plus_settings():
            valid = False
        
        # Validate trending analysis
        if not self.validate_trending_settings():
            valid = False
        
        # Validate minimum thresholds
        if not self.validate_minimum_thresholds():
            valid = False
        
        return valid
    
    def validate_thresholds(self):
        """
        Validate threshold logic and relationships.
        
        Returns:
            bool: True if thresholds are valid
        """
        thresholds = self.config.get('thresholds', {})
        efficiency = thresholds.get('efficiency_threshold', 0.15)
        success = thresholds.get('success_threshold', 0.3)
        green = thresholds.get('green_threshold', -0.1)
        yellow = thresholds.get('yellow_threshold', 0.1)
        red = thresholds.get('red_threshold', 0.3)
        
        valid = True
        
        # Efficiency should be <= success (stricter <= lenient)
        if efficiency > success:
            print(f"[ERROR] efficiency_threshold ({efficiency}) should be <= success_threshold ({success})")
            print(f"[HELP] Efficiency uses stricter criteria than general success")
            print(f"[HELP] Suggested: efficiency_threshold = {success} or lower")
            valid = False
        
        # Color thresholds should be in order
        if not (green < yellow < red):
            print(f"[ERROR] Color thresholds must be in order: green < yellow < red")
            print(f"[HELP] Current: green={green}, yellow={yellow}, red={red}")
            print(f"[HELP] Suggested: green=-0.1, yellow=0.1, red=0.3")
            valid = False
        
        # Red threshold should align with success threshold
        if abs(red - success) > 0.01:
            print(f"[WARN] red_threshold ({red}) should match success_threshold ({success}) for consistency")
            print(f"[HELP] Projects marked as 'red' should align with 'failed' projects")
        
        return valid
    
    def validate_data_source(self):
        """
        Validate data source configuration.
        
        Returns:
            bool: True if data source is valid
        """
        files = self.config.get('files', {})
        excel_file = files.get('excel_file', '')
        
        # Check if it's a file path or database connection
        if excel_file.lower().startswith(('server=', 'data source=', 'driver=')):
            print(f"[INFO] Database connection detected: {excel_file[:50]}...")
            print(f"[INFO] Skipping file existence check for database connection")
            return True
        
        # Validate Excel file
        if not excel_file:
            print(f"[ERROR] excel_file not specified in config")
            print(f"[HELP] Set 'excel_file' to your Excel file path or database connection string")
            return False
        
        if not excel_file.endswith(('.xlsx', '.xls')):
            print(f"[WARN] excel_file '{excel_file}' doesn't appear to be an Excel file")
            print(f"[HELP] Expected .xlsx or .xls extension, or database connection string")
        
        return True
    
    def validate_date_filtering(self):
        """
        Validate date filtering configuration.
        
        Returns:
            bool: True if date filtering is valid
        """
        filtering = self.config.get('project_filtering', {})
        
        if not filtering.get('enable_date_filter', False):
            return True  # Skip validation if disabled
        
        filter_type = filtering.get('filter_type', 'days')
        days_back = filtering.get('days_from_today', 90)
        specific_date = filtering.get('specific_date', '')
        
        valid = True
        
        # Validate days_from_today
        if filter_type == 'days':
            if days_back <= 0:
                print(f"[ERROR] days_from_today ({days_back}) must be positive")
                print(f"[HELP] Use positive number of days to look back (e.g., 90 for last 90 days)")
                valid = False
            elif days_back > 3650:  # 10 years
                print(f"[WARN] days_from_today ({days_back}) is very large (>10 years)")
                print(f"[HELP] Consider if you really need {days_back/365:.1f} years of data")
        
        # Validate specific_date
        elif filter_type == 'date':
            try:
                from datetime import datetime
                cutoff_date = datetime.strptime(specific_date, '%Y-%m-%d')
                if cutoff_date > datetime.now():
                    print(f"[ERROR] specific_date ({specific_date}) cannot be in the future")
                    print(f"[HELP] Use a past date or switch to 'days' filter type")
                    valid = False
            except ValueError:
                print(f"[ERROR] specific_date '{specific_date}' is not in YYYY-MM-DD format")
                print(f"[HELP] Use format like '2024-01-01'")
                valid = False
        
        return valid
    
    def validate_das_plus_settings(self):
        """
        Validate DAS+ analysis configuration.
        
        Returns:
            bool: True if DAS+ settings are valid
        """
        das_config = self.config.get('das_plus_analysis', {})
        
        if not das_config.get('enable_das_plus', False):
            return True  # Skip validation if disabled
        
        review_min = das_config.get('review_das_min', 0.3)
        review_max = das_config.get('review_das_max', 0.9)
        sample_n = das_config.get('sample_projects_per_consultant', 2)
        year_offset = das_config.get('current_year_offset', 0)
        
        valid = True
        
        # Validate DAS+ range
        if not (0.0 <= review_min <= 1.0):
            print(f"[ERROR] review_das_min ({review_min}) must be between 0.0 and 1.0")
            print(f"[HELP] DAS+ scores range from 0.0 (worst) to 1.0 (perfect)")
            valid = False
        
        if not (0.0 <= review_max <= 1.0):
            print(f"[ERROR] review_das_max ({review_max}) must be between 0.0 and 1.0")
            print(f"[HELP] DAS+ scores range from 0.0 (worst) to 1.0 (perfect)")
            valid = False
        
        if review_min >= review_max:
            print(f"[ERROR] review_das_min ({review_min}) must be < review_das_max ({review_max})")
            print(f"[HELP] Review range should capture middle-performing projects")
            print(f"[HELP] Suggested: min=0.3, max=0.9")
            valid = False
        
        # Validate sample count
        if sample_n <= 0:
            print(f"[ERROR] sample_projects_per_consultant ({sample_n}) must be positive")
            print(f"[HELP] Use 1-5 projects per consultant for review")
            valid = False
        
        # Year offset validation is handled in filter_data_by_year method
        
        return valid
    
    def validate_trending_settings(self):
        """
        Validate trending analysis configuration.
        
        Returns:
            bool: True if trending settings are valid
        """
        trending = self.config.get('trending_analysis', {})
        
        if not trending.get('enable_trending', False):
            return True  # Skip validation if disabled
        
        input_files = trending.get('input_files', [])
        
        if len(input_files) < 2:
            print(f"[ERROR] Trending analysis requires at least 2 input files")
            print(f"[HELP] Add more files to 'input_files' array for comparison")
            return False
        
        # Check file format (allow database connections)
        for file_config in input_files:
            file_path = file_config.get('file', '')
            period = file_config.get('period', '')
            
            if not file_path:
                print(f"[ERROR] Missing 'file' in trending input_files")
                print(f"[HELP] Each entry needs 'file' and 'period' fields")
                return False
            
            if not period:
                print(f"[ERROR] Missing 'period' for file {file_path}")
                print(f"[HELP] Each entry needs 'file' and 'period' fields")
                return False
        
        return True
    
    def validate_minimum_thresholds(self):
        """
        Validate minimum threshold settings.
        
        Returns:
            bool: True if minimum thresholds are valid
        """
        valid = True
        
        # Client analysis thresholds
        client_config = self.config.get('client_analysis', {})
        if client_config.get('enable_client_analysis', False):
            min_projects = client_config.get('min_projects_threshold', 3)
            if min_projects <= 0:
                print(f"[ERROR] min_projects_threshold ({min_projects}) must be positive")
                print(f"[HELP] Use 2-5 minimum projects for statistical relevance")
                valid = False
        
        # DAS+ analysis thresholds
        das_config = self.config.get('das_plus_analysis', {})
        if das_config.get('enable_das_plus', False):
            min_review = das_config.get('min_projects_for_review', 3)
            if min_review <= 0:
                print(f"[ERROR] min_projects_for_review ({min_review}) must be positive")
                print(f"[HELP] Use 2-5 minimum projects for consultant inclusion")
                valid = False
        
        # Inconsistent project check
        inconsistent_config = self.config.get('inconsistent_project_check', {})
        old_days = inconsistent_config.get('old_project_days', 730)
        if old_days <= 0:
            print(f"[ERROR] old_project_days ({old_days}) must be positive")
            print(f"[HELP] Use 365 (1 year), 730 (2 years), etc.")
            valid = False
        
        return valid
        
    def load_data(self):
        """
        Load and validate all input data files.
        
        Process:
        1. Load main Excel file with project data
        2. Combine duplicate job numbers intelligently
        3. Check for inconsistent project statuses
        4. Apply date filtering if enabled
        5. Load optional filter files (engineers, exclusions, SAs)
        
        Updates:
        - self.data: Main project DataFrame
        - self.engineers_list: Tracked consultants
        - self.exclusions: Project exclusions
        - self.solution_architects_list: Tracked SAs
        """
        try:
            # Load main Excel file with correct header row (configured row number)
            self.data = pd.read_excel(self.excel_file, header=self.config['files']['header_row'])
            print(f"[OK] Loaded {len(self.data)} records from {self.excel_file}")
            
            # Combine duplicate job numbers with intelligent field merging
            # This handles cases where same project appears multiple times
            self.data = self.combine_duplicate_jobs()
            print(f"[OK] After combining duplicates: {len(self.data)} records")
            
            # Check for data quality issues before filtering
            # Identifies projects needing status cleanup
            self.check_inconsistent_status_projects()
            
            # Store original data before filtering for customer analysis
            self.original_data = self.data.copy()
            
            # Apply date filtering if enabled in configuration
            # Excludes old closed projects based on end dates
            if self.config.get('project_filtering', {}).get('enable_date_filter', False):
                self.data = self.apply_date_filter()
                print(f"[OK] After date filtering: {len(self.data)} records")
            
            # Load engineers list if provided (optional filtering)
            # This file contains consultant names to focus analysis on
            if self.engineers_file and os.path.exists(self.engineers_file):
                with open(self.engineers_file, 'r') as f:
                    # Normalize names to handle spacing inconsistencies
                    self.engineers_list = [self.normalize_name(line.strip()) for line in f if line.strip()]
                print(f"[OK] Loaded {len(self.engineers_list)} engineers from {self.engineers_file}")
            else:
                print("[WARN] Engineers_PM.txt not found - will analyze all consultants in Resources Engaged")
                
            # Load exclusions if provided (optional filtering)
            # CSV format: Consultant,Project pairs to exclude from analysis
            if self.exclude_file and os.path.exists(self.exclude_file):
                exclude_df = pd.read_csv(self.exclude_file)
                self.exclusions = [(row['Consultant'], row['Project']) for _, row in exclude_df.iterrows()]
                print(f"[OK] Loaded {len(self.exclusions)} exclusions from {self.exclude_file}")
            else:
                print("[WARN] Exclude.csv not found - no exclusions applied")
                
            # Load solution architects list if provided (optional filtering)
            # Empty file means analyze all SAs found in data
            if self.solution_architects_file and os.path.exists(self.solution_architects_file):
                with open(self.solution_architects_file, 'r') as f:
                    self.solution_architects_list = [self.normalize_name(line.strip()) for line in f if line.strip()]
                if self.solution_architects_list:
                    print(f"[OK] Loaded {len(self.solution_architects_list)} solution architects from {self.solution_architects_file}")
                else:
                    print(f"[INFO] Solution architects file is empty - will analyze all solution architects")
            else:
                print("[WARN] Solution_Architects.txt not found - will analyze all solution architects")
                
        except Exception as e:
            print(f"[ERROR] Error loading data: {e}")
            sys.exit(1)
    
    def normalize_name(self, name):
        """
        Normalize consultant names to handle formatting inconsistencies.
        
        Common issues:
        - Double spaces between first/last names
        - Leading/trailing whitespace
        - Mixed case variations
        
        Args:
            name (str): Raw name from data
            
        Returns:
            str: Normalized name with single spaces
        """
        if pd.isna(name):
            return ""
        # Split on any whitespace and rejoin with single spaces
        return " ".join(str(name).split())
    
    def extract_consultants_from_resources(self, resources_text):
        """
        Extract consultant names from Resources Engaged field.
        
        Handles multiple formats:
        - Comma-separated: "John Doe, Jane Smith"
        - Semicolon-separated: "John Doe; Jane Smith"
        - Pipe-separated: "John Doe | Jane Smith"
        - Newline-separated: "John Doe\nJane Smith"
        - Single consultant: "John Doe"
        
        Args:
            resources_text (str): Raw resources field from Excel
            
        Returns:
            list: List of normalized consultant names
        """
        if pd.isna(resources_text):
            return []
        
        # Try common delimiters in order of preference
        consultants = []
        for delimiter in [',', ';', '|', '\n']:
            if delimiter in str(resources_text):
                parts = str(resources_text).split(delimiter)
                consultants.extend([self.normalize_name(part) for part in parts])
                break
        else:
            # No delimiter found, treat as single consultant
            consultants = [self.normalize_name(resources_text)]
        
        # Filter out empty names
        return [c for c in consultants if c]
    
    def is_excluded(self, consultant, project):
        """
        Check if consultant-project pair should be excluded from analysis.
        
        Uses exclusions loaded from Exclude.csv file.
        Names are normalized before comparison to handle formatting differences.
        
        Args:
            consultant (str): Consultant name to check
            project (str): Project identifier to check
            
        Returns:
            bool: True if this combination should be excluded
        """
        return any(self.normalize_name(consultant) == self.normalize_name(exc_consultant) and 
                  str(project) == str(exc_project) for exc_consultant, exc_project in self.exclusions)
    
    def calculate_efficiency_score(self, consultant_data):
        """
        Calculate efficiency score for a consultant using strict threshold.
        
        Efficiency scoring uses a stricter threshold (15% by default) compared
        to general success rate (30% by default). This creates a higher bar
        for "efficient" performance vs just "successful" completion.
        
        Logic:
        1. Only count projects with valid budget data
        2. Calculate variance: (actual - budget) / budget
        3. Project is "efficient" if variance <= efficiency_threshold
        4. Return percentage of efficient projects
        
        Args:
            consultant_data (DataFrame): Projects for this consultant
            
        Returns:
            float: Efficiency score as percentage (0-100)
        """
        if len(consultant_data) == 0:
            return 0
        
        successful_projects = 0
        total_projects_with_budget = 0
        
        # Evaluate each project for efficiency
        for _, project in consultant_data.iterrows():
            # Convert hours to numeric, defaulting to 0 if invalid
            actual_hours = pd.to_numeric(project.get('Total Hrs Posted', 0), errors='coerce') or 0
            budgeted_hours = pd.to_numeric(project.get('Budget Hrs', 0), errors='coerce') or 0
            
            # All projects in consultant_data already have valid budget data due to filtering
            total_projects_with_budget += 1
            # Calculate percentage variance from budget
            variance = (actual_hours - budgeted_hours) / budgeted_hours
            # Check if within efficiency threshold (stricter than success threshold)
            if variance <= self.config['thresholds']['efficiency_threshold']:
                successful_projects += 1
        
        # Avoid division by zero
        if total_projects_with_budget == 0:
            return 0
            
        # Return efficiency as percentage
        success_ratio = successful_projects / total_projects_with_budget
        return success_ratio * 100
    
    def analyze_consultants(self):
        """
        Main consultant analysis function.
        
        Process:
        1. Auto-detect relevant columns in Excel data
        2. Process each row to extract consultant assignments
        3. Group projects by consultant (handling duplicates)
        4. Calculate comprehensive metrics for each consultant
        5. Trigger solution architect analysis
        
        Metrics calculated:
        - Efficiency score (strict 15% threshold)
        - Success ratio (lenient 30% threshold) 
        - Composite score (efficiency + volume bonus)
        - Project counts and hour totals
        """
        print("\n[INFO] Analyzing consultant performance...")
        
        # Auto-detect relevant columns by searching for keywords
        # This handles variations in column naming across different Excel files
        resources_col = None      # "Resources Engaged" or similar
        job_number_col = None     # "Job Number" or similar
        actual_hours_col = None   # "Total Hrs Posted" or similar
        budgeted_hours_col = None # "Budget Hrs" or similar
        
        for col in self.data.columns:
            col_lower = str(col).lower()
            # Look for resources/consultants column
            if 'resource' in col_lower and 'engaged' in col_lower:
                resources_col = col
            # Look for project identifier column
            elif 'job' in col_lower and 'number' in col_lower:
                job_number_col = col
            # Look for actual hours column (multiple possible names)
            elif ('actual' in col_lower or 'total' in col_lower or 'posted' in col_lower) and ('hour' in col_lower or 'hrs' in col_lower):
                actual_hours_col = col
            # Look for budget hours column
            elif 'budget' in col_lower and ('hour' in col_lower or 'hrs' in col_lower):
                budgeted_hours_col = col
        
        # Validate that we found the essential columns
        if not resources_col:
            print("[ERROR] Could not find 'Resources Engaged' column")
            print("Available columns:")
            for i, col in enumerate(self.data.columns):
                print(f"  {i}: {col}")
            return
        
        print(f"[OK] Using columns: Resources='{resources_col}', Job='{job_number_col}', Actual='{actual_hours_col}', Budget='{budgeted_hours_col}'")
        
        # Process each row to build consultant-project relationships
        consultant_projects = {}  # Dict: consultant_name -> list of projects
        
        for idx, row in self.data.iterrows():
            # Extract data from current row
            resources = row.get(resources_col, "")
            job_number = row.get(job_number_col, f"Project_{idx}")
            
            # Check if project has valid budget data - exclude if not
            budgeted_hours = pd.to_numeric(row.get(budgeted_hours_col, 0), errors='coerce') or 0
            if budgeted_hours <= 0:
                continue  # Skip projects without valid budget data
            
            # Exclude cancelled projects
            project_status = str(row.get('Project Status', '')).lower()
            if 'cancel' in project_status:
                continue  # Skip cancelled projects
            
            # Parse consultants from resources field (handles comma-separated)
            consultants = self.extract_consultants_from_resources(resources)
            
            # Process each consultant found in this project
            for consultant in consultants:
                # Apply engineer filtering if list provided
                if self.engineers_list and self.normalize_name(consultant) not in self.engineers_list:
                    continue
                
                # Apply exclusions if any defined
                if self.is_excluded(consultant, job_number):
                    continue
                
                # Initialize consultant's project list if first time seeing them
                if consultant not in consultant_projects:
                    consultant_projects[consultant] = []
                
                # Add this project to consultant's portfolio (only valid budget projects)
                consultant_projects[consultant].append({
                    'Job Number': job_number,
                    'Actual Hours': pd.to_numeric(row.get(actual_hours_col, 0), errors='coerce') or 0,
                    'Budgeted Hours': budgeted_hours,
                    'Project Data': row  # Keep full row for detailed exports
                })
        
        # Calculate comprehensive metrics for each consultant
        for consultant, projects in consultant_projects.items():
            # Handle duplicate projects (same consultant on same project multiple times)
            unique_projects = {}  # Dict: job_number -> project_data
            total_hours = 0
            
            # Deduplicate projects while summing hours
            for project in projects:
                job_num = project['Job Number']
                if job_num not in unique_projects:
                    unique_projects[job_num] = project
                # Always add hours (handles cases where consultant appears multiple times)
                total_hours += project['Actual Hours']
            
            # Calculate efficiency score using strict threshold (15%)
            projects_df = pd.DataFrame([p['Project Data'] for p in unique_projects.values()])
            efficiency_score = self.calculate_efficiency_score(projects_df)
            
            # Calculate success ratio using lenient threshold (30%)
            # This creates two different performance measures
            successful_projects = 0
            projects_with_budget = 0
            for project in unique_projects.values():
                actual = project['Actual Hours']
                budgeted = project['Budgeted Hours']
                # All projects already have valid budget data due to filtering
                projects_with_budget += 1
                # Use success threshold (more lenient than efficiency threshold)
                if (actual - budgeted) / budgeted <= self.config['thresholds']['success_threshold']:
                    successful_projects += 1
            
            success_ratio = (successful_projects / projects_with_budget) * 100 if projects_with_budget > 0 else 0
            
            # Calculate composite score (efficiency + volume bonus)
            # This rewards both accuracy and volume of work
            hours_ratio = total_hours / self.config['scoring']['hours_per_bonus_point']
            hours_bonus = min(hours_ratio, self.config['scoring']['max_hours_multiplier']) * self.config['scoring']['bonus_points_per_1000_hours']
            composite_score = efficiency_score + hours_bonus
            
            # Check for on-hold projects
            on_hold_projects = 0
            for project in unique_projects.values():
                project_status = str(project['Project Data'].get('Project Status', '')).lower()
                if 'hold' in project_status:
                    on_hold_projects += 1
            
            # Store all calculated metrics
            self.consultant_metrics[consultant] = {
                'unique_projects': len(unique_projects),
                'total_hours': total_hours,
                'efficiency_score': efficiency_score,
                'composite_score': composite_score,
                'success_ratio': success_ratio,
                'projects_within_budget': successful_projects,
                'projects_over_budget': len(unique_projects) - successful_projects,
                'projects_on_hold': on_hold_projects
            }
        
        print(f"[OK] Analyzed {len(self.consultant_metrics)} consultants")
        
        # Analyze Solution Architects
        self.analyze_solution_architects()
    
    def analyze_solution_architects(self):
        """
        Analyze Solution Architect performance based on projects they designed.
        
        Key differences from consultant analysis:
        - SAs are measured on project success rates, not individual efficiency
        - Volume (hours sold/designed) is heavily weighted in ranking
        - Composite scoring balances success rate with business impact
        - Hours are split among multiple SAs on same project to avoid double-counting
        
        Process:
        1. Find Solution Architect column in data
        2. Parse comma-separated SA assignments
        3. Split project hours among multiple SAs
        4. Calculate success rates and volume metrics
        5. Generate composite scores for ranking
        """
        print("\n[INFO] Analyzing Solution Architect performance...")
        
        # Auto-detect Solution Architect column
        sa_col = None
        for col in self.data.columns:
            if 'solution' in str(col).lower() and 'architect' in str(col).lower():
                sa_col = col
                break
        
        if not sa_col:
            print("[WARN] No 'Solution Architect' column found - skipping SA analysis")
            return
        
        print(f"[OK] Using Solution Architect column: '{sa_col}'")
        
        # Process each project, handling multiple SAs per project
        sa_projects = {}  # Dict: sa_name -> list of projects
        
        for idx, row in self.data.iterrows():
            sa_field = str(row.get(sa_col, ""))
            # Skip rows with no SA assignment
            if not sa_field or sa_field.lower() in ['nan', 'none', '']:
                continue
            
            # Parse comma-separated SA names (e.g., "John Smith, Jane Doe")
            sa_names = [self.normalize_name(name.strip()) for name in sa_field.split(',')]
            sa_names = [name for name in sa_names if name]  # Remove empty names
            
            if not sa_names:
                continue
            
            # Extract project hours
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            
            # Check if project has valid budget data and apply exclusions
            if budgeted_hours > 0:
                # Exclude cancelled projects
                project_status = str(row.get('Project Status', '')).lower()
                if 'cancel' in project_status:
                    continue  # Skip cancelled projects
                # Check exclusions for each SA
                valid_sas = []
                for sa_name in sa_names:
                    # Filter by SA list if provided
                    if self.solution_architects_list and sa_name not in self.solution_architects_list:
                        continue
                    # Apply exclusions
                    if self.is_excluded(sa_name, row.get('Job Number', f'Project_{idx}')):
                        continue
                    valid_sas.append(sa_name)
                
                if not valid_sas:
                    continue  # Skip if no valid SAs after filtering
                
                # Split hours among valid SAs to avoid double counting
                split_actual = actual_hours / len(valid_sas)
                split_budgeted = budgeted_hours / len(valid_sas)
                # Variance calculated on full project (same for all SAs)
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                
                # Assign split hours to each valid SA
                for sa_name in valid_sas:
                    # Initialize SA's project list if first time
                    if sa_name not in sa_projects:
                        sa_projects[sa_name] = []
                    
                    # Add project with split hours to SA's portfolio
                    sa_projects[sa_name].append({
                        'variance': variance,
                        'actual_hours': split_actual,    # Split hours for volume calculation
                        'budgeted_hours': split_budgeted, # Split hours for volume calculation
                        'project_data': row
                    })
        
        # Calculate comprehensive SA metrics
        for sa_name, projects in sa_projects.items():
            total_projects = len(projects)
            # Count successful projects using success threshold (30%)
            successful_projects = sum(1 for p in projects if p['variance'] <= self.config['thresholds']['success_threshold'])
            # Sum split hours for volume metrics
            total_budgeted_hours = sum(p['budgeted_hours'] for p in projects)
            total_actual_hours = sum(p['actual_hours'] for p in projects)
            
            # Calculate success rate percentage
            success_rate = (successful_projects / total_projects) * 100 if total_projects > 0 else 0
            
            # Calculate composite score: success_rate × volume_multiplier
            # This rewards both accuracy and business impact (hours sold)
            volume_multiplier = min(total_budgeted_hours / self.config['solution_architect_scoring']['hours_per_multiplier'], 
                                  self.config['solution_architect_scoring']['max_volume_multiplier'])
            composite_score = success_rate * volume_multiplier
            
            # Store all SA metrics
            self.sa_metrics[sa_name] = {
                'total_projects': total_projects,
                'successful_projects': successful_projects,
                'failed_projects': total_projects - successful_projects,
                'success_rate': success_rate,
                'total_budgeted_hours': total_budgeted_hours,
                'total_actual_hours': total_actual_hours,
                'composite_score': composite_score
            }
        
        print(f"[OK] Analyzed {len(self.sa_metrics)} solution architects")
    
    def run_advanced_analytics(self):
        """
        Run advanced analytics suite on tracked projects only.
        
        Three main components:
        1. Predictive Analytics - Success rate forecasting by project size
        2. Anomaly Detection - Statistical outlier identification
        3. Risk Assessment - Multi-factor risk scoring
        
        All analytics focus only on projects involving tracked consultants
        or solution architects, not the entire company dataset.
        """
        print("\n" + "="*80)
        print("ADVANCED ANALYTICS")
        print("="*80)
        
        # Run predictive analytics if enabled
        if self.config['advanced_analytics'].get('enable_predictive', False):
            self.predictive_analysis()
        
        # Run anomaly detection if enabled
        if self.config['advanced_analytics'].get('enable_anomaly_detection', False):
            self.anomaly_detection()
        
        # Run risk assessment if enabled
        if self.config['advanced_analytics'].get('enable_risk_assessment', False):
            self.risk_assessment()
    
    def predictive_analysis(self):
        """
        Predict project success rates based on historical patterns.
        
        Analysis approach:
        - Categorize projects by budget size (small/medium/large)
        - Calculate historical success rates for each category
        - Provide predictions for future project planning
        - Generate consultant performance forecasts
        
        Only analyzes projects involving tracked team members.
        """
        print("\nPREDICTIVE ANALYTICS (Tracked Projects Only)")
        print("-" * 40)
        
        # Initialize project categorization buckets
        small_projects = []   # < 100 hours budgeted
        medium_projects = []  # 100-500 hours budgeted
        large_projects = []   # > 500 hours budgeted
        processed_projects = set()  # Track unique projects to avoid duplicates
        
        for idx, row in self.data.iterrows():
            # Only analyze projects involving tracked consultants or SAs
            if not self.is_tracked_project(row):
                continue
            
            # Get unique job number to avoid double counting
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
                
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                success = variance <= self.config['thresholds']['success_threshold']
                
                if budgeted_hours < 100:
                    small_projects.append(success)
                elif budgeted_hours <= 500:
                    medium_projects.append(success)
                else:
                    large_projects.append(success)
        
        # Calculate success rates by size
        small_rate = (sum(small_projects) / len(small_projects)) * 100 if small_projects else 0
        medium_rate = (sum(medium_projects) / len(medium_projects)) * 100 if medium_projects else 0
        large_rate = (sum(large_projects) / len(large_projects)) * 100 if large_projects else 0
        
        print(f"Success Rate Predictions by Project Size:")
        print(f"Small Projects (<100 hrs):   {small_rate:.1f}% ({len(small_projects)} projects)")
        print(f"Medium Projects (100-500):   {medium_rate:.1f}% ({len(medium_projects)} projects)")
        print(f"Large Projects (>500 hrs):   {large_rate:.1f}% ({len(large_projects)} projects)")
        
        # Consultant success prediction
        print(f"\nConsultant Performance Predictions:")
        sorted_consultants = sorted(self.consultant_metrics.items(), key=lambda x: x[1]['efficiency_score'], reverse=True)
        for consultant, metrics in sorted_consultants[:3]:
            predicted_success = min(95, metrics['efficiency_score'] + 5)  # Cap at 95%
            print(f"{consultant}: {predicted_success:.1f}% predicted success on new projects")
    
    def anomaly_detection(self):
        """
        Detect projects with statistically unusual performance.
        
        Statistical approach:
        - Calculate variance for all tracked projects
        - Compute mean and standard deviation of variances
        - Flag projects beyond configured threshold (default: 2 std devs)
        - Identify both extreme overruns and underruns
        
        Helps identify:
        - Severely underestimated projects
        - Scope creep situations
        - Exceptionally efficient executions
        - Data quality issues
        """
        print("\nANOMALY DETECTION (Tracked Projects Only)")
        print("-" * 40)
        
        variances = []  # List of all project variances for statistical analysis
        anomalies = []  # List of detected anomalous projects
        processed_projects = set()  # Avoid duplicate project analysis
        
        for idx, row in self.data.iterrows():
            # Only analyze projects involving tracked consultants or SAs
            if not self.is_tracked_project(row):
                continue
            
            # Get unique job number to avoid double counting
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
                
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                variances.append(variance)
                
                # Check for anomalies (beyond threshold standard deviations)
                if len(variances) > 10:  # Need some data for stats
                    mean_var = np.mean(variances)
                    std_var = np.std(variances)
                    threshold = self.config['advanced_analytics']['anomaly_threshold']
                    
                    if abs(variance - mean_var) > threshold * std_var:
                        current_job = row.get('Job Number', f'Project_{idx}')
                        anomalies.append({
                            'project': current_job,
                            'variance': variance * 100,
                            'budgeted': budgeted_hours,
                            'actual': actual_hours
                        })
        
        print(f"Detected {len(anomalies)} anomalous projects:")
        for anomaly in anomalies[-5:]:  # Show last 5
            # Get description for better identification
            job_desc = ""
            customer = ""
            # Find the row data for this project to get description
            for idx, row in self.data.iterrows():
                current_job = row.get('Job Number', f'Project_{idx}')
                if current_job == anomaly['project']:
                    job_desc = row.get('Job Description', '')
                    customer = row.get('Customer', '')
                    break
            
            # Build description with both customer and job description
            desc_parts = []
            if customer and str(customer).lower() not in ['nan', 'none', '']:
                desc_parts.append(f"Customer: {customer}")
            if job_desc and str(job_desc).lower() not in ['nan', 'none', '']:
                desc_parts.append(f"Desc: {job_desc}")
            
            desc_text = f" - {' | '.join(desc_parts)}" if desc_parts else ""
            print(f"  {anomaly['project']}{desc_text}: {anomaly['variance']:+.1f}% variance ({anomaly['budgeted']:.0f}→{anomaly['actual']:.0f} hrs)")
    
    def risk_assessment(self):
        """
        Identify high-risk projects using multi-factor scoring.
        
        Risk factors and scoring:
        - Large Budget (>500 hrs): +2 points
        - High Overrun (>50%): +3 points  
        - Complex Resourcing (>3 people): +1 point
        
        Risk score interpretation:
        - Score 3: Moderate risk - monitor closely
        - Score 4: High risk - intervention needed
        - Score 5+: Critical risk - escalate immediately
        
        Helps with:
        - Project portfolio management
        - Resource allocation decisions
        - Early warning system
        - Pattern recognition
        """
        print("\nRISK ASSESSMENT (Tracked Projects Only)")
        print("-" * 40)
        
        high_risk_projects = []  # Projects with risk score >= 3
        risk_patterns = {'large_budget': 0, 'new_consultant': 0, 'complex_resources': 0}  # Pattern counters
        processed_projects = set()  # Avoid duplicate analysis
        
        # Find resources column
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        for idx, row in self.data.iterrows():
            # Only analyze projects involving tracked consultants or SAs
            if not self.is_tracked_project(row):
                continue
            
            # Get unique job number to avoid double counting
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
                
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                risk_score = 0
                risk_factors = []
                
                # Risk factor 1: Large budget
                if budgeted_hours > self.config['advanced_analytics']['risk_threshold_hours']:
                    risk_score += 2
                    risk_factors.append('Large Budget')
                    risk_patterns['large_budget'] += 1
                
                # Risk factor 2: High variance threshold
                if variance > self.config['advanced_analytics']['risk_threshold_variance']:
                    risk_score += 3
                    risk_factors.append('High Overrun')
                
                # Risk factor 3: Multiple resources (complexity)
                if resources_col:
                    resources = str(row.get(resources_col, ''))
                    resource_count = len([r for r in resources.split(',') if r.strip()])
                    if resource_count > 3:
                        risk_score += 1
                        risk_factors.append('Complex Resourcing')
                        risk_patterns['complex_resources'] += 1
                
                if risk_score >= 3:  # High risk threshold
                    high_risk_projects.append({
                        'project': job_number,
                        'risk_score': risk_score,
                        'factors': risk_factors,
                        'variance': variance * 100,
                        'hours': budgeted_hours
                    })
        
        print(f"Identified {len(high_risk_projects)} high-risk projects:")
        for project in sorted(high_risk_projects, key=lambda x: x['risk_score'], reverse=True)[:5]:
            # Get description for better identification
            job_desc = ""
            customer = ""
            # Find the row data for this project to get description
            for idx, row in self.data.iterrows():
                current_job = row.get('Job Number', f'Project_{idx}')
                if current_job == project['project']:
                    job_desc = row.get('Job Description', '')
                    customer = row.get('Customer', '')
                    break
            
            # Build description with both customer and job description
            desc_parts = []
            if customer and str(customer).lower() not in ['nan', 'none', '']:
                desc_parts.append(f"Customer: {customer}")
            if job_desc and str(job_desc).lower() not in ['nan', 'none', '']:
                desc_parts.append(f"Desc: {job_desc}")
            
            desc_text = f" - {' | '.join(desc_parts)}" if desc_parts else ""
            factors_str = ', '.join(project['factors'])
            print(f"  {project['project']}: Risk Score {project['risk_score']}{desc_text} ({factors_str})")
        
        print(f"\nRisk Pattern Analysis:")
        print(f"Large Budget Projects (>{self.config['advanced_analytics']['risk_threshold_hours']} hrs): {risk_patterns['large_budget']}")
        print(f"Complex Resource Projects (>3 resources): {risk_patterns['complex_resources']}")
        
        # Export detailed analytics data
        self.export_advanced_analytics_details()
    
    def is_tracked_project(self, row):
        """
        Check if project involves tracked consultants or solution architects.
        
        A project is "tracked" if it involves:
        1. Any consultant from Engineers_PM_actual.txt, OR
        2. Any solution architect from Solution_Architects.txt (or any SA if list is empty)
        
        This filtering ensures advanced analytics focus on relevant projects
        rather than analyzing the entire company dataset.
        
        Args:
            row (Series): DataFrame row representing a project
            
        Returns:
            bool: True if project should be included in analytics
        """
        # Check for tracked consultants in Resources Engaged field
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        if resources_col:
            resources = row.get(resources_col, "")
            consultants = self.extract_consultants_from_resources(resources)
            has_tracked_consultant = any(
                self.normalize_name(consultant) in self.engineers_list 
                for consultant in consultants
            )
            if has_tracked_consultant:
                return True
        
        # Check for tracked solution architects
        sa_col = None
        for col in self.data.columns:
            if 'solution' in str(col).lower() and 'architect' in str(col).lower():
                sa_col = col
                break
        
        if sa_col:
            sa_field = str(row.get(sa_col, ""))
            if sa_field and sa_field.lower() not in ['nan', 'none', '']:
                sa_names = [self.normalize_name(name.strip()) for name in sa_field.split(',')]
                # If SA list is empty, include all SAs; otherwise filter by list
                has_tracked_sa = any(
                    not self.solution_architects_list or sa_name in self.solution_architects_list
                    for sa_name in sa_names if sa_name
                )
                if has_tracked_sa:
                    return True
        
        return False
    
    def run_trending_analysis(self):
        """
        Run trending analysis across multiple time periods.
        
        Process:
        1. Load and analyze each configured Excel file
        2. Calculate metrics for each period
        3. Compare metrics across periods
        4. Identify trends and changes
        5. Export trending report
        """
        print("\n" + "="*80)
        print("TRENDING ANALYSIS")
        print("="*80)
        
        trending_config = self.config['trending_analysis']
        input_files = trending_config['input_files']
        comparison_metrics = trending_config['comparison_metrics']
        
        # Store current state
        original_file = self.excel_file
        original_data = self.data
        
        period_metrics = {}  # Dict: period -> consultant_metrics
        
        # Process each period file
        for file_config in input_files:
            file_path = file_config['file']
            period_name = file_config['period']
            
            if not os.path.exists(file_path):
                print(f"[WARN] Trending file not found: {file_path} - skipping {period_name}")
                continue
            
            print(f"[INFO] Processing {period_name}: {file_path}")
            
            # Temporarily switch to this file
            self.excel_file = file_path
            self.consultant_metrics = {}
            self.sa_metrics = {}
            
            try:
                # Load and analyze this period's data
                self.load_data()
                self.analyze_consultants()
                
                # Store metrics for this period
                period_metrics[period_name] = {
                    'consultant_metrics': self.consultant_metrics.copy(),
                    'sa_metrics': self.sa_metrics.copy()
                }
                
                print(f"[OK] {period_name}: {len(self.consultant_metrics)} consultants, {len(self.sa_metrics)} SAs")
                
            except Exception as e:
                print(f"[ERROR] Failed to process {period_name}: {e}")
                continue
        
        # Restore original state
        self.excel_file = original_file
        self.data = original_data
        
        if len(period_metrics) < 2:
            print("[WARN] Need at least 2 periods for trending analysis")
            return
        
        # Generate trending analysis
        self.analyze_trends(period_metrics, comparison_metrics)
        self.export_trending_analysis(period_metrics, comparison_metrics)
    
    def analyze_trends(self, period_metrics, comparison_metrics):
        """
        Analyze trends across periods and display results.
        
        Args:
            period_metrics (dict): Metrics for each period
            comparison_metrics (list): Metrics to compare
        """
        periods = list(period_metrics.keys())
        print(f"\nTRENDING ANALYSIS RESULTS")
        print(f"Periods: {' → '.join(periods)}")
        print("-" * 60)
        
        # Get all consultants across all periods
        all_consultants = set()
        for period_data in period_metrics.values():
            all_consultants.update(period_data['consultant_metrics'].keys())
        
        # Analyze consultant trends
        print(f"\nCONSULTANT TRENDS")
        print(f"{'Consultant':<25} {'Metric':<15} {'Trend':<20} {'Change'}")
        print("-" * 75)
        
        trending_data = []
        
        for consultant in sorted(all_consultants):
            for metric in comparison_metrics:
                values = []
                for period in periods:
                    if consultant in period_metrics[period]['consultant_metrics']:
                        value = period_metrics[period]['consultant_metrics'][consultant].get(metric, 0)
                        values.append(value)
                    else:
                        values.append(None)
                
                # Calculate trend
                valid_values = [v for v in values if v is not None]
                if len(valid_values) >= 2:
                    first_val = valid_values[0]
                    last_val = valid_values[-1]
                    change = last_val - first_val
                    change_pct = (change / first_val * 100) if first_val != 0 else 0
                    
                    trend_str = f"{first_val:.1f} → {last_val:.1f}"
                    change_str = f"{change_pct:+.1f}%"
                    
                    print(f"{consultant:<25} {metric:<15} {trend_str:<20} {change_str}")
                    
                    trending_data.append({
                        'consultant': consultant,
                        'metric': metric,
                        'first_value': first_val,
                        'last_value': last_val,
                        'change': change,
                        'change_percent': change_pct
                    })
        
        # Show top improvers and decliners
        if trending_data:
            efficiency_trends = [t for t in trending_data if t['metric'] == 'efficiency_score']
            if efficiency_trends:
                print(f"\nTOP PERFORMERS")
                top_improvers = sorted(efficiency_trends, key=lambda x: x['change_percent'], reverse=True)[:3]
                for trend in top_improvers:
                    print(f"📈 Most Improved: {trend['consultant']} ({trend['change_percent']:+.1f}% efficiency change)")
                
                top_decliners = sorted(efficiency_trends, key=lambda x: x['change_percent'])[:3]
                for trend in top_decliners:
                    if trend['change_percent'] < -5:  # Only show significant declines
                        print(f"📉 Needs Attention: {trend['consultant']} ({trend['change_percent']:+.1f}% efficiency change)")
    
    def export_trending_analysis(self, period_metrics, comparison_metrics):
        """
        Export trending analysis to Excel file.
        
        Args:
            period_metrics (dict): Metrics for each period
            comparison_metrics (list): Metrics to compare
        """
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'trending_analysis_{timestamp}.xlsx'
        
        periods = list(period_metrics.keys())
        all_consultants = set()
        for period_data in period_metrics.values():
            all_consultants.update(period_data['consultant_metrics'].keys())
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary trends tab
            trend_data = []
            for consultant in sorted(all_consultants):
                row = {'Consultant': consultant}
                
                for metric in comparison_metrics:
                    for period in periods:
                        if consultant in period_metrics[period]['consultant_metrics']:
                            value = period_metrics[period]['consultant_metrics'][consultant].get(metric, 0)
                            row[f"{metric}_{period}"] = value
                        else:
                            row[f"{metric}_{period}"] = None
                    
                    # Calculate trend
                    values = [row[f"{metric}_{period}"] for period in periods if row[f"{metric}_{period}"] is not None]
                    if len(values) >= 2:
                        change = values[-1] - values[0]
                        change_pct = (change / values[0] * 100) if values[0] != 0 else 0
                        row[f"{metric}_change"] = change
                        row[f"{metric}_change_pct"] = change_pct
                
                trend_data.append(row)
            
            trend_df = pd.DataFrame(trend_data)
            trend_df.to_excel(writer, sheet_name='Trending_Summary', index=False)
            
            # Individual period tabs
            for period, metrics in period_metrics.items():
                period_data = []
                for consultant, consultant_metrics in metrics['consultant_metrics'].items():
                    row = {'Consultant': consultant}
                    row.update(consultant_metrics)
                    period_data.append(row)
                
                if period_data:
                    period_df = pd.DataFrame(period_data)
                    sheet_name = period.replace(' ', '_')[:31]  # Excel sheet name limit
                    period_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"[EXPORT] Trending analysis exported to: {output_file}")
    
    def run_client_analysis(self):
        """
        Analyze performance by client and consultant-client combinations.
        
        Process:
        1. Group projects by client
        2. Calculate success rates and efficiency by client
        3. Analyze consultant performance with specific clients
        4. Identify best/worst client relationships
        5. Export detailed client analysis
        """
        print("\n" + "="*80)
        print("CLIENT PERFORMANCE ANALYSIS")
        print("="*80)
        
        client_config = self.config['client_analysis']
        min_projects = client_config.get('min_projects_threshold', 3)
        track_consultant_client = client_config.get('track_consultant_client_performance', True)
        
        # Find customer column
        client_col = None
        for col in self.data.columns:
            col_lower = str(col).lower()
            if 'customer' in col_lower:
                client_col = col
                break
        
        if not client_col:
            print("[WARN] No 'Customer' column found - skipping client analysis")
            return
        
        print(f"[OK] Using customer column: '{client_col}'")
        
        # Analyze client performance
        client_metrics = self.analyze_client_performance(client_col, min_projects)
        
        # Analyze consultant-client combinations if enabled
        consultant_client_metrics = {}
        if track_consultant_client:
            consultant_client_metrics = self.analyze_consultant_client_performance(client_col, min_projects)
        
        # Display results
        self.display_client_analysis(client_metrics, consultant_client_metrics, min_projects)
        
        # Export results
        self.export_client_analysis(client_metrics, consultant_client_metrics)
    
    def analyze_client_performance(self, client_col, min_projects):
        """
        Analyze overall performance by client.
        
        Args:
            client_col (str): Name of client column
            min_projects (int): Minimum projects to include client
            
        Returns:
            dict: Client performance metrics
        """
        client_projects = {}
        
        # Group projects by customer (using ALL data, not filtered)
        data_to_use = self.original_data if self.original_data is not None else self.data
        for idx, row in data_to_use.iterrows():
            client = self.normalize_name(str(row.get(client_col, "")))
            if not client or client.lower() in ['nan', 'none', '']:
                continue
            
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            
            # Only process projects with valid budget data and not cancelled
            if budgeted_hours > 0:
                # Exclude cancelled projects
                project_status = str(row.get('Project Status', '')).lower()
                if 'cancel' in project_status:
                    continue  # Skip cancelled projects
                
                if client not in client_projects:
                    client_projects[client] = []
                
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                client_projects[client].append({
                    'actual_hours': actual_hours,
                    'budgeted_hours': budgeted_hours,
                    'variance': variance,
                    'job_number': row.get('Job Number', f'Project_{idx}'),
                    'project_data': row
                })
        
        # Calculate metrics for each customer
        client_metrics = {}
        for client, projects in client_projects.items():
            if len(projects) >= min_projects:
                total_projects = len(projects)
                successful_projects = sum(1 for p in projects if p['variance'] <= self.config['thresholds']['success_threshold'])
                efficient_projects = sum(1 for p in projects if p['variance'] <= self.config['thresholds']['efficiency_threshold'])
                
                total_budgeted = sum(p['budgeted_hours'] for p in projects)
                total_actual = sum(p['actual_hours'] for p in projects)
                avg_variance = sum(p['variance'] for p in projects) / total_projects
                
                client_metrics[client] = {
                    'total_projects': total_projects,
                    'successful_projects': successful_projects,
                    'efficient_projects': efficient_projects,
                    'success_rate': (successful_projects / total_projects) * 100,
                    'efficiency_rate': (efficient_projects / total_projects) * 100,
                    'total_budgeted_hours': total_budgeted,
                    'total_actual_hours': total_actual,
                    'avg_variance_pct': avg_variance * 100,
                    'projects': projects
                }
        
        return client_metrics
    
    def analyze_consultant_client_performance(self, client_col, min_projects):
        """
        Analyze consultant performance with specific clients.
        
        Args:
            client_col (str): Name of client column
            min_projects (int): Minimum projects for analysis
            
        Returns:
            dict: Consultant-client performance metrics
        """
        consultant_client_projects = {}
        
        # Find resources column
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        if not resources_col:
            return {}
        
        # Group projects by consultant-customer combination (using ALL data, not filtered)
        data_to_use = self.original_data if self.original_data is not None else self.data
        for idx, row in data_to_use.iterrows():
            client = self.normalize_name(str(row.get(client_col, "")))
            if not client or client.lower() in ['nan', 'none', '']:
                continue
            
            resources = row.get(resources_col, "")
            consultants = self.extract_consultants_from_resources(resources)
            
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            
            # Only process projects with valid budget data and not cancelled
            if budgeted_hours > 0:
                # Exclude cancelled projects
                project_status = str(row.get('Project Status', '')).lower()
                if 'cancel' in project_status:
                    continue  # Skip cancelled projects
                
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                
                for consultant in consultants:
                    # Filter by engineers list if provided
                    if self.engineers_list and self.normalize_name(consultant) not in self.engineers_list:
                        continue
                    
                    # Apply exclusions
                    if self.is_excluded(consultant, row.get('Job Number', f'Project_{idx}')):
                        continue
                    
                    combo_key = f"{consultant}|{client}"
                    if combo_key not in consultant_client_projects:
                        consultant_client_projects[combo_key] = []
                    
                    consultant_client_projects[combo_key].append({
                        'actual_hours': actual_hours,
                        'budgeted_hours': budgeted_hours,
                        'variance': variance,
                        'job_number': row.get('Job Number', f'Project_{idx}')
                    })
        
        # Calculate metrics for consultant-customer combinations
        consultant_client_metrics = {}
        for combo_key, projects in consultant_client_projects.items():
            if len(projects) >= min_projects:
                consultant, client = combo_key.split('|')
                total_projects = len(projects)
                successful_projects = sum(1 for p in projects if p['variance'] <= self.config['thresholds']['success_threshold'])
                efficient_projects = sum(1 for p in projects if p['variance'] <= self.config['thresholds']['efficiency_threshold'])
                
                avg_variance = sum(p['variance'] for p in projects) / total_projects
                
                consultant_client_metrics[combo_key] = {
                    'consultant': consultant,
                    'client': client,
                    'total_projects': total_projects,
                    'success_rate': (successful_projects / total_projects) * 100,
                    'efficiency_rate': (efficient_projects / total_projects) * 100,
                    'avg_variance_pct': avg_variance * 100
                }
        
        return consultant_client_metrics
    
    def display_client_analysis(self, client_metrics, consultant_client_metrics, min_projects):
        """
        Display client analysis results to console.
        
        Args:
            client_metrics (dict): Client performance data
            consultant_client_metrics (dict): Consultant-client performance data
            min_projects (int): Minimum projects threshold
        """
        if client_metrics:
            print(f"\nCUSTOMER PERFORMANCE SUMMARY (Min {min_projects} projects)")
            print(f"{'Customer':<30} {'Projects':<8} {'Success':<8} {'Efficiency':<10} {'Avg Variance'}")
            print("-" * 75)
            
            # Sort by success rate
            sorted_clients = sorted(client_metrics.items(), key=lambda x: x[1]['success_rate'], reverse=True)
            for client, metrics in sorted_clients:
                success_str = f"{metrics['success_rate']:.1f}%"
                efficiency_str = f"{metrics['efficiency_rate']:.1f}%"
                variance_str = f"{metrics['avg_variance_pct']:+.1f}%"
                print(f"{client:<30} {metrics['total_projects']:<8} "
                      f"{success_str:<8} {efficiency_str:<10} "
                      f"{variance_str}")
            
            # Show best and worst clients
            if len(sorted_clients) >= 2:
                best_client = sorted_clients[0]
                worst_client = sorted_clients[-1]
                print(f"\n🏆 Best Customer: {best_client[0]} ({best_client[1]['success_rate']:.1f}% success rate)")
                print(f"🚨 Most Challenging: {worst_client[0]} ({worst_client[1]['success_rate']:.1f}% success rate)")
        
        if consultant_client_metrics:
            print(f"\nTOP CONSULTANT-CUSTOMER COMBINATIONS")
            print(f"{'Consultant':<20} {'Customer':<25} {'Projects':<8} {'Success':<8} {'Efficiency'}")
            print("-" * 80)
            
            # Sort by efficiency rate
            sorted_combos = sorted(consultant_client_metrics.items(), 
                                 key=lambda x: x[1]['efficiency_rate'], reverse=True)
            for combo_key, metrics in sorted_combos[:10]:  # Top 10
                success_str = f"{metrics['success_rate']:.1f}%"
                efficiency_str = f"{metrics['efficiency_rate']:.1f}%"
                print(f"{metrics['consultant']:<20} {metrics['client']:<25} "
                      f"{metrics['total_projects']:<8} {success_str:<8} "
                      f"{efficiency_str}")
    
    def export_client_analysis(self, client_metrics, consultant_client_metrics):
        """
        Export client analysis to Excel file.
        
        Args:
            client_metrics (dict): Client performance data
            consultant_client_metrics (dict): Consultant-client performance data
        """
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'client_analysis_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Customer summary tab
            if client_metrics:
                client_data = []
                for client, metrics in client_metrics.items():
                    client_data.append({
                        'Customer': client,
                        'Total_Projects': metrics['total_projects'],
                        'Successful_Projects': metrics['successful_projects'],
                        'Efficient_Projects': metrics['efficient_projects'],
                        'Success_Rate': metrics['success_rate'],
                        'Efficiency_Rate': metrics['efficiency_rate'],
                        'Total_Budgeted_Hours': metrics['total_budgeted_hours'],
                        'Total_Actual_Hours': metrics['total_actual_hours'],
                        'Avg_Variance_Pct': metrics['avg_variance_pct']
                    })
                
                client_df = pd.DataFrame(client_data)
                client_df = client_df.sort_values('Success_Rate', ascending=False)
                client_df.to_excel(writer, sheet_name='Customer_Summary', index=False)
            
            # Consultant-customer combinations tab
            if consultant_client_metrics:
                combo_data = []
                for combo_key, metrics in consultant_client_metrics.items():
                    combo_data.append({
                        'Consultant': metrics['consultant'],
                        'Customer': metrics['client'],
                        'Total_Projects': metrics['total_projects'],
                        'Success_Rate': metrics['success_rate'],
                        'Efficiency_Rate': metrics['efficiency_rate'],
                        'Avg_Variance_Pct': metrics['avg_variance_pct']
                    })
                
                combo_df = pd.DataFrame(combo_data)
                combo_df = combo_df.sort_values('Efficiency_Rate', ascending=False)
                combo_df.to_excel(writer, sheet_name='Consultant_Customer_Performance', index=False)
        
        print(f"[EXPORT] Customer analysis exported to: {output_file}")
    
    def run_das_plus_analysis(self):
        """
        Run DAS+ (Delivery Accuracy Score Plus) analysis.
        
        DAS+ accounts for project completion percentage and provides
        a more nuanced view of consultant performance by considering
        both budget accuracy and delivery progress.
        """
        print("\n" + "="*80)
        print("DAS+ ANALYSIS (Delivery Accuracy Score Plus)")
        print("="*80)
        
        das_config = self.config['das_plus_analysis']
        min_projects = das_config.get('min_projects_for_review', 3)
        review_min = das_config.get('review_das_min', 0.3)
        review_max = das_config.get('review_das_max', 0.9)
        sample_n = das_config.get('sample_projects_per_consultant', 2)
        
        # Calculate DAS+ for all projects
        data_with_das = self.calculate_das_plus_scores()
        
        if data_with_das.empty:
            print("[WARN] No projects with valid DAS+ scores found")
            return
        
        # Generate consultant DAS+ summary (all projects)
        das_summary_all = self.generate_das_summary(data_with_das, min_projects, "All Time")
        
        # Generate current year DAS+ summary
        current_year_data = self.filter_data_by_year(data_with_das)
        das_summary_year = self.generate_das_summary(current_year_data, min_projects, "Current Year")
        
        # Generate date-filtered DAS+ summary (for review projects)
        filtered_data = self.filter_data_by_date_range(data_with_das)
        das_summary_filtered = self.generate_das_summary(filtered_data, min_projects, "Date Filtered")
        
        # Display results
        self.display_das_analysis(das_summary_all, das_summary_year, das_summary_filtered)
        
        # Generate review projects from date-filtered data only
        review_projects = self.get_review_projects(filtered_data, das_summary_filtered, review_min, review_max, sample_n)
        
        # Export combined results
        self.export_combined_das_analysis(das_summary_all, das_summary_year, das_summary_filtered, review_projects)
    
    def calculate_das_plus_scores(self):
        """
        Calculate DAS+ scores for all projects.
        
        Returns:
            DataFrame: Projects with DAS+ scores
        """
        import numpy as np
        
        # Use original data for DAS+ (not filtered)
        data_to_use = self.original_data if self.original_data is not None else self.data
        data_copy = data_to_use.copy()
        
        def compute_das_plus(row):
            """Improved DAS that accounts for completed/closed projects."""
            try:
                budget = float(row.get('Budget Hrs', 0))
                actual = float(row.get('Total Hrs Posted', 0))
                complete_pct = float(row.get('Project Complete %', 0))
                status = str(row.get('Project Status', '')).lower()
        
                if budget <= 0 or np.isnan(complete_pct):
                    return None
        
                # Normalize completion to 100% if project is closed or 95%+
                completion_ratio = 1.0 if ('closed' in status or complete_pct >= 95) else complete_pct / 100
                budget_ratio = actual / budget
                das = 1 - abs(budget_ratio - completion_ratio)
                return round(max(0.0, min(1.0, das)), 4)
            except:
                return None
        
        # Calculate DAS+ scores
        data_copy['DAS_Plus'] = data_copy.apply(compute_das_plus, axis=1)
        
        # Explode consultants from Resources Engaged
        data_copy['Consultants'] = data_copy['Resources Engaged'].astype(str).str.split(',')
        data_exploded = data_copy.explode('Consultants')
        data_exploded['Consultants'] = data_exploded['Consultants'].str.strip()
        
        # Filter by tracked consultants if provided
        if self.engineers_list:
            data_exploded = data_exploded[data_exploded['Consultants'].apply(
                lambda x: self.normalize_name(x) in self.engineers_list
            )]
        
        # Apply exclusions for DAS+ analysis
        filtered_rows = []
        for idx, row in data_exploded.iterrows():
            consultant = row['Consultants']
            job_number = row.get('Job Number', f'Project_{idx}')
            
            # Skip if excluded
            if self.is_excluded(consultant, job_number):
                continue
            
            filtered_rows.append(row)
        
        if filtered_rows:
            data_exploded = pd.DataFrame(filtered_rows)
        else:
            data_exploded = pd.DataFrame()
        
        # Remove rows with invalid DAS+ scores
        data_exploded = data_exploded.dropna(subset=['DAS_Plus', 'Consultants'])
        
        return data_exploded
    
    def filter_data_by_year(self, data_with_das):
        """
        Filter data to current year only.
        
        Args:
            data_with_das (DataFrame): Projects with DAS+ scores
            
        Returns:
            DataFrame: Current year projects only
        """
        from datetime import datetime
        year_offset = self.config.get('das_plus_analysis', {}).get('current_year_offset', 0)
        current_year = datetime.now().year + year_offset
        
        # Validate year offset - cannot analyze future years
        if year_offset > 0:
            print(f"[ERROR] Cannot analyze future years (current_year_offset: {year_offset})")
            print(f"[HELP] DAS+ analysis can only be performed on past or current year data.")
            print(f"[HELP] Please set 'current_year_offset' in config.json to:")
            print(f"[HELP]   0  = Current year ({datetime.now().year})")
            print(f"[HELP]   -1 = Previous year ({datetime.now().year - 1})")
            print(f"[HELP]   -2 = Two years ago ({datetime.now().year - 2}), etc.")
            return pd.DataFrame()  # Return empty DataFrame
        
        # Find date column
        date_col = None
        for col in data_with_das.columns:
            if 'end' in str(col).lower() and 'date' in str(col).lower():
                date_col = col
                break
        
        if not date_col:
            return data_with_das  # Return all if no date column found
        
        # Filter by current year
        filtered_data = []
        for idx, row in data_with_das.iterrows():
            end_date_str = row.get(date_col, '')
            if pd.notna(end_date_str):
                try:
                    date_str = str(end_date_str).split()[0]
                    for date_format in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y']:
                        try:
                            end_date = datetime.strptime(date_str, date_format)
                            if end_date.year == current_year:
                                filtered_data.append(row)
                            break
                        except ValueError:
                            continue
                except:
                    continue
        
        return pd.DataFrame(filtered_data) if filtered_data else pd.DataFrame()
    
    def filter_data_by_date_range(self, data_with_das):
        """
        Filter data using the same date filter as main analysis.
        
        Args:
            data_with_das (DataFrame): Projects with DAS+ scores
            
        Returns:
            DataFrame: Date-filtered projects
        """
        from datetime import datetime, timedelta
        
        # Use same filtering logic as main analysis
        filter_config = self.config.get('project_filtering', {})
        if not filter_config.get('enable_date_filter', False):
            return data_with_das  # Return all if filtering disabled
        
        # Determine cutoff date
        if filter_config['filter_type'] == 'days':
            cutoff_date = datetime.now() - timedelta(days=filter_config['days_from_today'])
        else:
            cutoff_date = datetime.strptime(filter_config['specific_date'], '%Y-%m-%d')
        
        # Find relevant columns
        status_col = None
        end_date_col = None
        
        for col in data_with_das.columns:
            col_lower = str(col).lower()
            if 'project' in col_lower and 'status' in col_lower:
                status_col = col
            elif 'end' in col_lower and 'date' in col_lower:
                end_date_col = col
        
        if not status_col or not end_date_col:
            return data_with_das
        
        filtered_data = []
        for idx, row in data_with_das.iterrows():
            include_project = True
            
            if filter_config['exclude_closed_before_date']:
                status = str(row.get(status_col, '')).lower()
                end_date_str = row.get(end_date_col, '')
                
                if status == 'closed' and pd.notna(end_date_str):
                    try:
                        date_str = str(end_date_str).split()[0]
                        end_date = None
                        for date_format in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y']:
                            try:
                                end_date = datetime.strptime(date_str, date_format)
                                break
                            except ValueError:
                                continue
                        
                        if end_date and end_date < cutoff_date:
                            include_project = False
                    except:
                        pass
            
            if include_project:
                filtered_data.append(row)
        
        return pd.DataFrame(filtered_data) if filtered_data else pd.DataFrame()
    
    def generate_das_summary(self, data_with_das, min_projects, period_name=""):
        """
        Generate consultant-level DAS+ summary.
        
        Args:
            data_with_das (DataFrame): Projects with DAS+ scores
            min_projects (int): Minimum projects for inclusion
            period_name (str): Name of the period for identification
            
        Returns:
            DataFrame: Consultant DAS+ summary
        """
        das_summary = (
            data_with_das.groupby('Consultants')
            .agg(
                Project_Count=('DAS_Plus', 'count'),
                Avg_DAS_Plus=('DAS_Plus', 'mean'),
                Median_DAS_Plus=('DAS_Plus', 'median'),
                Low_DAS_Count=('DAS_Plus', lambda x: (x < 0.75).sum()),
                High_DAS_Count=('DAS_Plus', lambda x: (x >= 0.85).sum())
            )
            .sort_values(by='Avg_DAS_Plus', ascending=False)
            .reset_index()
        )
        
        # Filter to reliable consultants
        das_summary = das_summary[das_summary['Project_Count'] >= min_projects]
        
        # Add period identifier
        das_summary['Period'] = period_name
        
        return das_summary
    
    def display_das_analysis(self, das_summary_all, das_summary_year, das_summary_filtered):
        """
        Display DAS+ analysis results for all periods.
        
        Args:
            das_summary_all (DataFrame): All-time DAS+ summary
            das_summary_year (DataFrame): Current year DAS+ summary
            das_summary_filtered (DataFrame): Date-filtered DAS+ summary
        """
        # Display All-Time Summary
        if not das_summary_all.empty:
            print(f"\nDAS+ CONSULTANT SUMMARY - ALL TIME")
            print(f"{'Consultant':<25} {'Projects':<8} {'Avg DAS+':<10} {'Median':<8} {'Low (<0.75)':<12} {'High (≥0.85)'}")
            print("-" * 85)
            
            for _, row in das_summary_all.iterrows():
                print(f"{row['Consultants']:<25} {row['Project_Count']:<8} "
                      f"{row['Avg_DAS_Plus']:<10.3f} {row['Median_DAS_Plus']:<8.3f} "
                      f"{row['Low_DAS_Count']:<12} {row['High_DAS_Count']}")
        
        # Display Current Year Summary
        if not das_summary_year.empty:
            print(f"\nDAS+ CONSULTANT SUMMARY - CURRENT YEAR")
            print(f"{'Consultant':<25} {'Projects':<8} {'Avg DAS+':<10} {'Median':<8} {'Low (<0.75)':<12} {'High (≥0.85)'}")
            print("-" * 85)
            
            for _, row in das_summary_year.iterrows():
                print(f"{row['Consultants']:<25} {row['Project_Count']:<8} "
                      f"{row['Avg_DAS_Plus']:<10.3f} {row['Median_DAS_Plus']:<8.3f} "
                      f"{row['Low_DAS_Count']:<12} {row['High_DAS_Count']}")
        
        # Display Date-Filtered Summary
        if not das_summary_filtered.empty:
            print(f"\nDAS+ CONSULTANT SUMMARY - DATE FILTERED (Recent Projects)")
            print(f"{'Consultant':<25} {'Projects':<8} {'Avg DAS+':<10} {'Median':<8} {'Low (<0.75)':<12} {'High (≥0.85)'}")
            print("-" * 85)
            
            for _, row in das_summary_filtered.iterrows():
                print(f"{row['Consultants']:<25} {row['Project_Count']:<8} "
                      f"{row['Avg_DAS_Plus']:<10.3f} {row['Median_DAS_Plus']:<8.3f} "
                      f"{row['Low_DAS_Count']:<12} {row['High_DAS_Count']}")
        
        # Show overall best performer
        if not das_summary_all.empty:
            best = das_summary_all.iloc[0]
            worst = das_summary_all.iloc[-1]
            print(f"\n🏆 Best Overall DAS+ Performance: {best['Consultants']} (All-time Avg: {best['Avg_DAS_Plus']:.3f})")
            print(f"⚠️ Needs Attention: {worst['Consultants']} (All-time Avg: {worst['Avg_DAS_Plus']:.3f})")
    
    def get_review_projects(self, data_with_das, das_summary, min_das, max_das, sample_n):
        """
        Pick sample projects per consultant for review from recent projects only.
        
        Args:
            data_with_das (DataFrame): Date-filtered projects with DAS+ scores
            das_summary (DataFrame): Consultant summary
            min_das (float): Minimum DAS+ for review
            max_das (float): Maximum DAS+ for review
            sample_n (int): Number of projects to sample per consultant
            
        Returns:
            DataFrame: Sample projects for review (recent projects only)
        """
        review_rows = []
        reliable_consultants = das_summary['Consultants'].tolist()
        
        for consultant in reliable_consultants:
            consultant_df = data_with_das[data_with_das['Consultants'] == consultant]
            
            # Limit to projects with DAS scores in given range
            consultant_df = consultant_df[
                consultant_df['DAS_Plus'].between(min_das, max_das)
            ]
            
            if not consultant_df.empty:
                # Sample projects for review
                review_sample = consultant_df.sample(
                    n=min(sample_n, len(consultant_df)), 
                    random_state=42
                )
                review_rows.append(review_sample)
        
        if review_rows:
            return pd.concat(review_rows, ignore_index=True)
        return pd.DataFrame()
    
    def export_combined_das_analysis(self, das_summary_all, das_summary_year, das_summary_filtered, review_projects):
        """
        Export combined DAS+ analysis results to single file.
        
        Args:
            das_summary_all (DataFrame): All-time DAS+ summary
            das_summary_year (DataFrame): Current year DAS+ summary
            das_summary_filtered (DataFrame): Date-filtered DAS+ summary
            review_projects (DataFrame): Projects for review
        """
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Export combined DAS+ analysis
        output_file = f'das_plus_analysis_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # All-time summary
            if not das_summary_all.empty:
                das_summary_all.to_excel(writer, sheet_name='All_Time_Summary', index=False)
            
            # Current year summary
            if not das_summary_year.empty:
                das_summary_year.to_excel(writer, sheet_name='Current_Year_Summary', index=False)
            
            # Date-filtered summary
            if not das_summary_filtered.empty:
                das_summary_filtered.to_excel(writer, sheet_name='Recent_Projects_Summary', index=False)
            
            # Review projects (from recent data only)
            if not review_projects.empty:
                # Remove unnamed columns for cleaner export
                review_clean = review_projects.loc[:, ~review_projects.columns.str.startswith('Unnamed')]
                # Move Consultants column to front for easier identification
                if 'Consultants' in review_clean.columns:
                    cols = ['Consultants'] + [col for col in review_clean.columns if col != 'Consultants']
                    review_clean = review_clean[cols]
                review_clean.to_excel(writer, sheet_name='Review_Projects', index=False)
        
        print(f"[EXPORT] Combined DAS+ analysis exported to: {output_file}")
        if review_projects.empty:
            print(f"[INFO] No recent projects found in review range (DAS+ {min_das}-{max_das})")
    
    def export_advanced_analytics_details(self):
        """Export detailed Advanced Analytics data to Excel with project breakdowns"""
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'advanced_analytics_details_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary and Explanations (first tab)
            self.export_analytics_summary(writer)
            
            # Analytics Report (console output)
            self.export_analytics_report(writer)
            
            # Predictive Analysis Projects
            self.export_predictive_projects(writer)
            
            # Anomaly Detection Projects
            self.export_anomaly_projects(writer)
            
            # Risk Assessment Projects
            self.export_risk_projects(writer)
        
        print(f"[EXPORT] Advanced Analytics details exported to: {output_file}")
    
    def export_predictive_projects(self, writer):
        """Export projects used in predictive analysis by size category"""
        small_projects = []
        medium_projects = []
        large_projects = []
        processed_projects = set()
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                success = variance <= self.config['thresholds']['success_threshold']
                
                project_data = {
                    'Job_Number': job_number,
                    'Budget_Hrs': budgeted_hours,
                    'Actual_Hrs': actual_hours,
                    'Variance_%': variance * 100,
                    'Success': success,
                    'Resources_Engaged': row.get('Resources Engaged', ''),
                    'Solution_Architect': row.get('Solution Architect', ''),
                    'Account_Manager': row.get('Account Manager', '')
                }
                
                if budgeted_hours < 100:
                    small_projects.append(project_data)
                elif budgeted_hours <= 500:
                    medium_projects.append(project_data)
                else:
                    large_projects.append(project_data)
        
        # Export each category
        if small_projects:
            pd.DataFrame(small_projects).to_excel(writer, sheet_name='Small_Projects', index=False)
        if medium_projects:
            pd.DataFrame(medium_projects).to_excel(writer, sheet_name='Medium_Projects', index=False)
        if large_projects:
            pd.DataFrame(large_projects).to_excel(writer, sheet_name='Large_Projects', index=False)
    
    def export_anomaly_projects(self, writer):
        """Export projects identified as anomalies"""
        variances = []
        anomaly_projects = []
        processed_projects = set()
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                variances.append(variance)
                
                # Check for anomalies
                if len(variances) > 10:
                    mean_var = np.mean(variances)
                    std_var = np.std(variances)
                    threshold = self.config['advanced_analytics']['anomaly_threshold']
                    
                    if abs(variance - mean_var) > threshold * std_var:
                        # Get both customer and job description for better identification
                        job_desc = row.get('Job Description', '')
                        customer = row.get('Customer', '')
                        
                        anomaly_projects.append({
                            'Job_Number': job_number,
                            'Customer': customer,
                            'Job_Description': job_desc,
                            'Budget_Hrs': budgeted_hours,
                            'Actual_Hrs': actual_hours,
                            'Variance_%': variance * 100,
                            'Deviation_Score': abs(variance - mean_var) / std_var,
                            'Resources_Engaged': row.get('Resources Engaged', ''),
                            'Solution_Architect': row.get('Solution Architect', ''),
                            'Account_Manager': row.get('Account Manager', ''),
                            'Project_Description': row.get('Project Description', ''),
                            'Customer': customer
                        })
        
        if anomaly_projects:
            pd.DataFrame(anomaly_projects).to_excel(writer, sheet_name='Anomaly_Projects', index=False)
    
    def export_risk_projects(self, writer):
        """Export high-risk projects with risk factors"""
        high_risk_projects = []
        processed_projects = set()
        
        # Find resources column
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                risk_score = 0
                risk_factors = []
                
                # Risk factor 1: Large budget
                if budgeted_hours > self.config['advanced_analytics']['risk_threshold_hours']:
                    risk_score += 2
                    risk_factors.append('Large Budget')
                
                # Risk factor 2: High variance threshold
                if variance > self.config['advanced_analytics']['risk_threshold_variance']:
                    risk_score += 3
                    risk_factors.append('High Overrun')
                
                # Risk factor 3: Multiple resources (complexity)
                if resources_col:
                    resources = str(row.get(resources_col, ''))
                    resource_count = len([r for r in resources.split(',') if r.strip()])
                    if resource_count > 3:
                        risk_score += 1
                        risk_factors.append('Complex Resourcing')
                
                if risk_score >= 3:  # High risk threshold
                    # Get both customer and job description for better identification
                    job_desc = row.get('Job Description', '')
                    customer = row.get('Customer', '')
                    
                    high_risk_projects.append({
                        'Job_Number': job_number,
                        'Risk_Score': risk_score,
                        'Customer': customer,
                        'Job_Description': job_desc,
                        'Risk_Factors': ', '.join(risk_factors),
                        'Budget_Hrs': budgeted_hours,
                        'Actual_Hrs': actual_hours,
                        'Variance_%': variance * 100,
                        'Resources_Engaged': row.get('Resources Engaged', ''),
                        'Solution_Architect': row.get('Solution Architect', ''),
                        'Account_Manager': row.get('Account Manager', ''),
                        'Project_Description': row.get('Project Description', ''),
                        'Customer': customer
                    })
        
        if high_risk_projects:
            # Sort by risk score
            high_risk_df = pd.DataFrame(high_risk_projects)
            high_risk_df = high_risk_df.sort_values('Risk_Score', ascending=False)
            high_risk_df.to_excel(writer, sheet_name='High_Risk_Projects', index=False)
    
    def export_analytics_summary(self, writer):
        """Export summary and explanations of Advanced Analytics"""
        summary_data = [
            ['ADVANCED ANALYTICS SUMMARY & EXPLANATIONS', ''],
            ['', ''],
            ['OVERVIEW', ''],
            ['This report analyzes projects involving tracked consultants and solution architects only.', ''],
            ['All calculations exclude company-wide projects without tracked team members.', ''],
            ['', ''],
            ['PREDICTIVE ANALYTICS', ''],
            ['Projects are categorized by budget size to predict success rates:', ''],
            ['• Small Projects: <100 hours budgeted', ''],
            ['• Medium Projects: 100-500 hours budgeted', ''],
            ['• Large Projects: >500 hours budgeted', ''],
            ['Success Rate = % of projects within 30% of budget', ''],
            ['', ''],
            ['ANOMALY DETECTION', ''],
            ['Statistical analysis identifies projects with unusual performance:', ''],
            ['• Uses standard deviation to find outliers', ''],
            ['• Deviation Score >2.0 = Anomaly (configurable)', ''],
            ['• Higher deviation = more unusual performance', ''],
            ['', ''],
            ['RISK ASSESSMENT', ''],
            ['Projects scored based on risk factors:', ''],
            ['Risk Score 3+ = High Risk Project', ''],
            ['', ''],
            ['Risk Factors & Points:', ''],
            ['• Large Budget (>500 hrs): +2 points', ''],
            ['• High Overrun (>50% over budget): +3 points', ''],
            ['• Complex Resourcing (>3 people): +1 point', ''],
            ['', ''],
            ['Risk Score Meanings:', ''],
            ['• Score 3: Moderate Risk - Monitor closely', ''],
            ['• Score 4: High Risk - Requires intervention', ''],
            ['• Score 5: Critical Risk - Immediate action needed', ''],
            ['• Score 6: Extreme Risk - Escalate to management', ''],
            ['', ''],
            ['VARIANCE EXPLANATIONS', ''],
            ['Variance shows how much actual hours differ from budget:', ''],
            ['', ''],
            ['Example: "+289.0% variance (36→141 hrs)"', ''],
            ['• Budgeted: 36 hours', ''],
            ['• Actual: 141 hours', ''],
            ['• Variance: +289% (project took 3.9x longer than planned)', ''],
            ['• This indicates major scope creep or estimation issues', ''],
            ['', ''],
            ['Variance Interpretation:', ''],
            ['• 0% = Exactly on budget (rare)', ''],
            ['• +10% = 10% over budget (acceptable)', ''],
            ['• +30% = 30% over budget (concerning)', ''],
            ['• +50% = 50% over budget (problematic)', ''],
            ['• +100% = Double the budget (major issue)', ''],
            ['• +200%+ = Triple+ the budget (critical failure)', ''],
            ['', ''],
            ['• -10% = 10% under budget (good efficiency)', ''],
            ['• -20% = 20% under budget (excellent efficiency)', ''],
            ['• -50% = 50% under budget (possible under-scoping)', ''],
            ['', ''],
            ['SUCCESS CRITERIA', ''],
            ['• Efficiency Threshold: 15% over budget', ''],
            ['• Success Threshold: 30% over budget', ''],
            ['• Projects >30% over = Failed projects', ''],
            ['• Projects ≤30% over = Successful projects', ''],
            ['', ''],
            ['COLOR CODING (in project tabs)', ''],
            ['• Green: >10% under budget (excellent)', ''],
            ['• Yellow: 10-30% over budget (acceptable)', ''],
            ['• Red: >30% over budget (failed)', ''],
            ['', ''],
            ['RECOMMENDATIONS', ''],
            ['• Review all Red projects for lessons learned', ''],
            ['• Investigate Anomaly projects for root causes', ''],
            ['• Monitor High Risk projects closely', ''],
            ['• Celebrate Green projects and share best practices', ''],
            ['• Use predictive data for better project planning', '']
        ]
        
        summary_df = pd.DataFrame(summary_data, columns=['Description', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary_Explanations', index=False)
    
    def export_analytics_report(self, writer):
        """Export Advanced Analytics console report to Excel"""
        report_data = []
        
        # Predictive Analytics Results
        small_projects = []
        medium_projects = []
        large_projects = []
        processed_projects = set()
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                success = variance <= self.config['thresholds']['success_threshold']
                
                if budgeted_hours < 100:
                    small_projects.append(success)
                elif budgeted_hours <= 500:
                    medium_projects.append(success)
                else:
                    large_projects.append(success)
        
        # Calculate success rates
        small_rate = (sum(small_projects) / len(small_projects)) * 100 if small_projects else 0
        medium_rate = (sum(medium_projects) / len(medium_projects)) * 100 if medium_projects else 0
        large_rate = (sum(large_projects) / len(large_projects)) * 100 if large_projects else 0
        
        # Anomaly Detection Results
        variances = []
        anomalies = []
        processed_projects = set()
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                variances.append(variance)
                
                if len(variances) > 10:
                    mean_var = np.mean(variances)
                    std_var = np.std(variances)
                    threshold = self.config['advanced_analytics']['anomaly_threshold']
                    
                    if abs(variance - mean_var) > threshold * std_var:
                        anomalies.append({
                            'project': job_number,
                            'variance': variance * 100,
                            'budgeted': budgeted_hours,
                            'actual': actual_hours
                        })
        
        # Risk Assessment Results
        high_risk_projects = []
        processed_projects = set()
        
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        for idx, row in self.data.iterrows():
            if not self.is_tracked_project(row):
                continue
            
            job_number = row.get('Job Number', f'Project_{idx}')
            if job_number in processed_projects:
                continue
            processed_projects.add(job_number)
            
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                risk_score = 0
                risk_factors = []
                
                if budgeted_hours > self.config['advanced_analytics']['risk_threshold_hours']:
                    risk_score += 2
                    risk_factors.append('Large Budget')
                
                if variance > self.config['advanced_analytics']['risk_threshold_variance']:
                    risk_score += 3
                    risk_factors.append('High Overrun')
                
                if resources_col:
                    resources = str(row.get(resources_col, ''))
                    resource_count = len([r for r in resources.split(',') if r.strip()])
                    if resource_count > 3:
                        risk_score += 1
                        risk_factors.append('Complex Resourcing')
                
                if risk_score >= 3:
                    high_risk_projects.append({
                        'project': job_number,
                        'risk_score': risk_score,
                        'factors': risk_factors,
                        'variance': variance * 100,
                        'hours': budgeted_hours
                    })
        
        # Build report data
        report_data = [
            ['ADVANCED ANALYTICS REPORT', ''],
            ['Generated from tracked projects only', ''],
            ['', ''],
            ['PREDICTIVE ANALYTICS', ''],
            ['Success Rate Predictions by Project Size:', ''],
            [f'Small Projects (<100 hrs)', f'{small_rate:.1f}% ({len(small_projects)} projects)'],
            [f'Medium Projects (100-500)', f'{medium_rate:.1f}% ({len(medium_projects)} projects)'],
            [f'Large Projects (>500 hrs)', f'{large_rate:.1f}% ({len(large_projects)} projects)'],
            ['', ''],
            ['Key Insight:', 'Larger projects tend to have lower success rates'],
            ['', ''],
            ['ANOMALY DETECTION', ''],
            [f'Detected {len(anomalies)} anomalous projects:', ''],
        ]
        
        # Add top 5 anomalies
        for anomaly in sorted(anomalies, key=lambda x: abs(x['variance']), reverse=True)[:5]:
            report_data.append([f"  {anomaly['project']}", f"{anomaly['variance']:+.1f}% variance ({anomaly['budgeted']:.0f}→{anomaly['actual']:.0f} hrs)"])
        
        report_data.extend([
            ['', ''],
            ['RISK ASSESSMENT', ''],
            [f'Identified {len(high_risk_projects)} high-risk projects:', ''],
        ])
        
        # Add top 5 high-risk projects
        for project in sorted(high_risk_projects, key=lambda x: x['risk_score'], reverse=True)[:5]:
            factors_str = ', '.join(project['factors'])
            report_data.append([f"  {project['project']}", f"Risk Score {project['risk_score']} ({factors_str})"])
        
        report_data.extend([
            ['', ''],
            ['RECOMMENDATIONS', ''],
            ['• Review anomalous projects for root causes', ''],
            ['• Monitor high-risk projects closely', ''],
            ['• Use predictive data for project planning', ''],
            ['• Focus on improving large project success rates', '']
        ])
        
        report_df = pd.DataFrame(report_data, columns=['Metric', 'Value'])
        report_df.to_excel(writer, sheet_name='Analytics_Report', index=False)
    
    def generate_report(self):
        """Generate comprehensive performance report"""
        if not self.consultant_metrics:
            print("[ERROR] No metrics to report")
            return
        
        print("\n" + "="*80)
        print("CONSULTANT PERFORMANCE METRICS REPORT")
        print("="*80)
        
        # Summary statistics
        total_consultants = len(self.consultant_metrics)
        total_hours = sum(m['total_hours'] for m in self.consultant_metrics.values())
        avg_efficiency = np.mean([m['efficiency_score'] for m in self.consultant_metrics.values()])
        total_projects = self.get_unique_total_projects()
        
        print(f"\nSUMMARY STATISTICS")
        print(f"{'Total Consultants:':<25} {total_consultants}")
        print(f"{'Total Hours Logged:':<25} {total_hours:,.1f}")
        print(f"{'Total Unique Projects:':<25} {total_projects}")
        print(f"{'Average Efficiency:':<25} {avg_efficiency:.1f}%")
        
        # Sort consultants by composite score
        sorted_consultants = sorted(self.consultant_metrics.items(), 
                                  key=lambda x: x[1]['composite_score'], reverse=True)
        
        # Top performers
        print(f"\nTOP PERFORMERS")
        if self.show_composite:
            print(f"{'Consultant':<30} {'Projects':<8} {'Hours':<8} {'Efficiency':<10} {'Composite':<10} {'Success'}")
            print("-" * 85)
            for consultant, metrics in sorted_consultants[:self.config['display']['top_performers_count']]:
                efficiency_str = f"{metrics['efficiency_score']:.1f}%"
                print(f"{consultant:<30} {metrics['unique_projects']:<8} "
                      f"{metrics['total_hours']:<8.0f} {efficiency_str:<10} "
                      f"{metrics['composite_score']:<10.1f} {metrics['success_ratio']:.1f}%")
        else:
            print(f"{'Consultant':<30} {'Projects':<8} {'Hours':<8} {'Efficiency':<10} {'Success Rate'}")
            print("-" * 75)
            for consultant, metrics in sorted_consultants[:self.config['display']['top_performers_count']]:
                efficiency_str = f"{metrics['efficiency_score']:.1f}%"
                print(f"{consultant:<30} {metrics['unique_projects']:<8} "
                      f"{metrics['total_hours']:<8.0f} {efficiency_str:<10} "
                      f"{metrics['success_ratio']:.1f}%")
        
        # Detailed metrics table
        print(f"\nDETAILED CONSULTANT METRICS")
        if self.show_composite:
            print(f"{'Consultant':<30} {'Projects':<8} {'Hours':<8} {'Efficiency':<10} {'Composite':<10} {'Success':<7} {'Over':<4} {'Hold'}")
            print("-" * 100)
            for consultant, metrics in sorted_consultants:
                hold_indicator = f"({metrics['projects_on_hold']})" if metrics['projects_on_hold'] > 0 else ""
                efficiency_str = f"{metrics['efficiency_score']:.1f}%"
                print(f"{consultant:<30} {metrics['unique_projects']:<8} "
                      f"{metrics['total_hours']:<8.0f} {efficiency_str:<10} "
                      f"{metrics['composite_score']:<10.1f} {metrics['projects_within_budget']:<7} {metrics['projects_over_budget']:<4} {hold_indicator}")
        else:
            print(f"{'Consultant':<30} {'Projects':<8} {'Hours':<8} {'Efficiency':<10} {'Success':<7} {'Over':<4} {'Hold'}")
            print("-" * 90)
            for consultant, metrics in sorted_consultants:
                hold_indicator = f"({metrics['projects_on_hold']})" if metrics['projects_on_hold'] > 0 else ""
                efficiency_str = f"{metrics['efficiency_score']:.1f}%"
                print(f"{consultant:<30} {metrics['unique_projects']:<8} "
                      f"{metrics['total_hours']:<8.0f} {efficiency_str:<10} "
                      f"{metrics['projects_within_budget']:<7} {metrics['projects_over_budget']:<4} {hold_indicator}")
        
        # Highlight special categories
        if sorted_consultants:
            print(f"\nSPECIAL RECOGNITIONS")
            
            # Highest composite score
            best_consultant = sorted_consultants[0]
            if self.show_composite:
                print(f"Consultant of the Year: {best_consultant[0]} (Composite: {best_consultant[1]['composite_score']:.1f}, Efficiency: {best_consultant[1]['efficiency_score']:.1f}%)")
            else:
                print(f"Consultant of the Year: {best_consultant[0]} ({best_consultant[1]['efficiency_score']:.1f}% efficiency)")
            
            # Most hours
            most_hours = max(self.consultant_metrics.items(), key=lambda x: x[1]['total_hours'])
            print(f"Most Hours Contributor: {most_hours[0]} ({most_hours[1]['total_hours']:.1f} hours)")
            
            # Most projects
            most_projects = max(self.consultant_metrics.items(), key=lambda x: x[1]['unique_projects'])
            print(f"Most Projects Engaged: {most_projects[0]} ({most_projects[1]['unique_projects']} projects)")
        
        # Practice Performance Analysis (unique projects with tracked consultants)
        practice_projects = self.get_unique_practice_projects()
        
        success_threshold_pct = int(self.config['thresholds']['success_threshold'] * 100)
        print(f"\nPRACTICE PERFORMANCE ANALYSIS (Tracked Consultants)")
        print(f"Projects within {success_threshold_pct}% of budget: {practice_projects['within']}")
        print(f"Projects over {success_threshold_pct}% of budget: {practice_projects['over']}")
        practice_total = practice_projects['within'] + practice_projects['over']
        if practice_total > 0:
            print(f"Practice success rate: {(practice_projects['within']/practice_total)*100:.1f}%")
        
        # Show excluded projects count
        excluded_count = practice_projects['excluded_count']
        if excluded_count > 0:
            print(f"Note: {excluded_count} projects excluded due to missing/invalid budget data")
        
        # Company Performance Analysis (all projects)
        self.analyze_company_performance()
        
        # Generate SA report if we have SA metrics
        if self.sa_metrics:
            self.generate_sa_report()
    
    def get_unique_practice_projects(self):
        """Get unique projects that involved any tracked consultant"""
        unique_projects = set()
        excluded_projects = []
        practice_within = 0
        practice_over = 0
        
        # Find resources column
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        if not resources_col:
            return {'within': 0, 'over': 0, 'excluded_count': 0}
        
        for idx, row in self.data.iterrows():
            resources = row.get(resources_col, "")
            consultants = self.extract_consultants_from_resources(resources)
            
            # Check if any tracked consultant is in this project
            has_tracked_consultant = any(
                self.normalize_name(consultant) in self.engineers_list 
                for consultant in consultants
            )
            
            if has_tracked_consultant:
                job_number = None
                for col in self.data.columns:
                    if 'job' in str(col).lower() and 'number' in str(col).lower():
                        job_number = row.get(col, f"Project_{idx}")
                        break
                
                if job_number not in unique_projects:
                    unique_projects.add(job_number)
                    
                    actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
                    budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
                    
                    if budgeted_hours > 0:
                        variance = (actual_hours - budgeted_hours) / budgeted_hours
                        if variance <= self.config['thresholds']['success_threshold']:
                            practice_within += 1
                        else:
                            practice_over += 1
                    else:
                        # Track excluded project
                        excluded_projects.append(row)
        
        # Export excluded projects if any
        if excluded_projects:
            self.export_excluded_projects(excluded_projects)
        

        
        return {'within': practice_within, 'over': practice_over, 'excluded_count': len(excluded_projects)}
    
    def get_unique_total_projects(self):
        """Get count of unique projects involving any tracked consultant"""
        unique_projects = set()
        
        # Find resources column
        resources_col = None
        for col in self.data.columns:
            if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                resources_col = col
                break
        
        if not resources_col:
            return 0
        
        for idx, row in self.data.iterrows():
            resources = row.get(resources_col, "")
            consultants = self.extract_consultants_from_resources(resources)
            
            # Check if any tracked consultant is in this project
            has_tracked_consultant = any(
                self.normalize_name(consultant) in self.engineers_list 
                for consultant in consultants
            )
            
            if has_tracked_consultant:
                job_number = None
                for col in self.data.columns:
                    if 'job' in str(col).lower() and 'number' in str(col).lower():
                        job_number = row.get(col, f"Project_{idx}")
                        break
                
                if job_number:
                    unique_projects.add(job_number)
        
        return len(unique_projects)
    
    def combine_duplicate_jobs(self):
        """
        Combine rows with same job number, intelligently merging data.
        
        Handles cases where the same project appears multiple times in the
        Excel file, which can happen due to:
        - Multiple phases or tasks
        - Data entry errors
        - Different reporting periods
        
        Merging strategy:
        - Numeric fields (hours, amounts): Sum values
        - Text fields (names, descriptions): Prefer non-N/A values, combine unique entries
        - Dates and other fields: Keep first non-null value
        
        Returns:
            DataFrame: Deduplicated data with combined values
        """
        # Find job number column for grouping
        job_col = None
        for col in self.data.columns:
            if 'job' in str(col).lower() and 'number' in str(col).lower():
                job_col = col
                break
        
        if not job_col:
            return self.data
        
        # Group by job number and process each group
        combined_data = []
        job_groups = self.data.groupby(job_col)
        
        for job_number, group in job_groups:
            if len(group) == 1:
                # Single entry, keep as is
                combined_data.append(group.iloc[0])
            else:
                # Multiple entries for same job, combine intelligently
                combined_row = group.iloc[0].copy()  # Start with first row
                
                # Sum numeric columns (hours, budget amounts)
                numeric_cols = ['Budget Hrs', 'Total Hrs Posted', 'Budget Amount', 'Actual Amount']
                for col in numeric_cols:
                    if col in group.columns:
                        # Convert to numeric and sum, handling invalid values
                        combined_row[col] = pd.to_numeric(group[col], errors='coerce').sum()
                
                # For text columns, prefer valid values over N/A
                text_cols = ['Account Manager', 'Resources Engaged', 'Solution Architect']
                for col in text_cols:
                    if col in group.columns:
                        # Get all non-null, non-N/A values
                        valid_values = group[col].dropna()
                        valid_values = valid_values[valid_values.str.upper() != 'N/A']
                        
                        if len(valid_values) > 0:
                            # Combine unique values with comma separation
                            unique_values = list(set(valid_values.astype(str)))
                            combined_row[col] = ', '.join(unique_values)
                        # else keep the original value (might be N/A)
                
                combined_data.append(combined_row)
        
        return pd.DataFrame(combined_data)
    
    def apply_date_filter(self):
        """
        Filter projects based on configured date criteria.
        
        Filtering logic:
        - Include all open projects (regardless of date)
        - Include closed projects within the date range
        - Exclude closed projects before the cutoff date
        
        Date calculation options:
        - Relative: X days from today (e.g., last 90 days)
        - Absolute: Specific cutoff date (e.g., 2025-01-01)
        
        Returns:
            DataFrame: Filtered project data
        """
        from datetime import datetime, timedelta
        
        filter_config = self.config['project_filtering']
        
        # Determine cutoff date (projects before this date will be excluded if closed)
        if filter_config['filter_type'] == 'days':
            cutoff_date = datetime.now() - timedelta(days=filter_config['days_from_today'])
        else:  # specific date
            cutoff_date = datetime.strptime(filter_config['specific_date'], '%Y-%m-%d')
        
        if self.show_composite:
            print(f"[DEBUG] Current date: {datetime.now().strftime('%Y-%m-%d')}")
            print(f"[DEBUG] Cutoff date: {cutoff_date.strftime('%Y-%m-%d')}")
            print(f"[DEBUG] Will exclude closed projects with end dates before: {cutoff_date.strftime('%Y-%m-%d')}")
        
        # Find relevant columns
        status_col = None
        end_date_col = None
        
        for col in self.data.columns:
            col_lower = str(col).lower()
            if 'status' in col_lower:
                status_col = col
            elif 'end' in col_lower and 'date' in col_lower:
                end_date_col = col
        
        if self.show_composite:
            print(f"[DEBUG] Available columns: {list(self.data.columns)}")
            print(f"[DEBUG] Found Status column: {status_col}")
            print(f"[DEBUG] Found End Date column: {end_date_col}")
        
        if not status_col or not end_date_col:
            print(f"[WARN] Could not find Project Status or End Date columns - skipping date filter")
            print(f"[WARN] Looking for columns containing 'status' and 'end date'")
            return self.data
        
        print(f"[OK] Using columns: Status='{status_col}', End Date='{end_date_col}'")
        print(f"[OK] Excluding closed projects before: {cutoff_date.strftime('%Y-%m-%d')}")
        
        filtered_data = []
        excluded_count = 0
        
        for idx, row in self.data.iterrows():
            include_project = True
            
            if filter_config['exclude_closed_before_date']:
                status = str(row.get(status_col, '')).lower()
                end_date_str = row.get(end_date_col, '')
                
                if status == 'closed' and pd.notna(end_date_str):
                    try:
                        # Try multiple date formats
                        end_date = None
                        date_str = str(end_date_str).split()[0]  # Remove time if present
                        
                        for date_format in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S']:
                            try:
                                end_date = datetime.strptime(date_str, date_format)
                                break
                            except ValueError:
                                continue
                        
                        if end_date:
                            if self.show_composite:
                                print(f"[DEBUG] Project {row.get('Job Number', 'Unknown')}: Status={status}, End Date={end_date.strftime('%Y-%m-%d')}, Cutoff={cutoff_date.strftime('%Y-%m-%d')}")
                            if end_date < cutoff_date:
                                include_project = False
                                excluded_count += 1
                                if self.show_composite:
                                    print(f"[DEBUG] → EXCLUDING (closed before cutoff)")
                            else:
                                if self.show_composite:
                                    print(f"[DEBUG] → INCLUDING (closed after cutoff)")
                        else:
                            if self.show_composite:
                                print(f"[DEBUG] Could not parse date: {end_date_str}")
                    except (ValueError, TypeError) as e:
                        # If date parsing fails, include the project
                        if self.show_composite:
                            print(f"[DEBUG] Date parsing failed for {row.get('Job Number', 'Unknown')}: {end_date_str} - {e}")
                else:
                    if self.show_composite:
                        print(f"[DEBUG] Project {row.get('Job Number', 'Unknown')}: Status={status} → INCLUDING (not closed)")
            
            if include_project:
                filtered_data.append(row)

        
        if excluded_count > 0:
            print(f"[OK] Excluded {excluded_count} closed projects with end dates before {cutoff_date.strftime('%Y-%m-%d')}")
        return pd.DataFrame(filtered_data)
    
    def analyze_company_performance(self):
        """Analyze overall company performance"""
        company_within = 0
        company_over = 0
        
        for idx, row in self.data.iterrows():
            actual_hours = pd.to_numeric(row.get('Total Hrs Posted', 0), errors='coerce') or 0
            budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
            
            if budgeted_hours > 0:
                variance = (actual_hours - budgeted_hours) / budgeted_hours
                if variance <= self.config['thresholds']['success_threshold']:
                    company_within += 1
                else:
                    company_over += 1
        
        company_total = company_within + company_over
        success_threshold_pct = int(self.config['thresholds']['success_threshold'] * 100)
        
        print(f"\nCOMPANY PERFORMANCE ANALYSIS (All Projects)")
        print(f"Projects within {success_threshold_pct}% of budget: {company_within}")
        print(f"Projects over {success_threshold_pct}% of budget: {company_over}")
        if company_total > 0:
            print(f"Company success rate: {(company_within/company_total)*100:.1f}%")
    
    def generate_sa_report(self):
        """Generate Solution Architect performance report"""
        print("\n" + "="*80)
        print("SOLUTION ARCHITECT PERFORMANCE REPORT")
        print("="*80)
        
        # Sort by composite score
        sorted_sas = sorted(self.sa_metrics.items(), key=lambda x: x[1]['composite_score'], reverse=True)
        
        print(f"\nSOLUTION ARCHITECT RANKINGS")
        if self.show_composite:
            print(f"{'Solution Architect':<30} {'Projects':<8} {'Hours':<8} {'Success':<8} {'Composite':<10} {'Variance'}")
            print("-" * 90)
            for sa_name, metrics in sorted_sas:
                variance = ((metrics['total_actual_hours'] - metrics['total_budgeted_hours']) / metrics['total_budgeted_hours']) * 100 if metrics['total_budgeted_hours'] > 0 else 0
                success_str = f"{metrics['success_rate']:.1f}%"
                variance_str = f"{variance:+.1f}%"
                print(f"{sa_name:<30} {metrics['total_projects']:<8} "
                      f"{metrics['total_budgeted_hours']:<8.0f} {success_str:<8} "
                      f"{metrics['composite_score']:<10.1f} {variance_str}")
        else:
            print(f"{'Solution Architect':<30} {'Projects':<8} {'Hours':<8} {'Success Rate':<12} {'Variance'}")
            print("-" * 85)
            for sa_name, metrics in sorted_sas:
                variance = ((metrics['total_actual_hours'] - metrics['total_budgeted_hours']) / metrics['total_budgeted_hours']) * 100 if metrics['total_budgeted_hours'] > 0 else 0
                success_str = f"{metrics['success_rate']:.1f}%"
                variance_str = f"{variance:+.1f}%"
                print(f"{sa_name:<30} {metrics['total_projects']:<8} "
                      f"{metrics['total_budgeted_hours']:<8.0f} {success_str:<12} "
                      f"{variance_str}")
        
        if sorted_sas:
            best_sa = sorted_sas[0]
            if self.show_composite:
                print(f"\nTop Solution Architect: {best_sa[0]} (Composite: {best_sa[1]['composite_score']:.1f}, {best_sa[1]['total_budgeted_hours']:.0f} hours, {best_sa[1]['success_rate']:.1f}% success)")
            else:
                print(f"\nTop Solution Architect: {best_sa[0]} ({best_sa[1]['total_budgeted_hours']:.0f} hours sold, {best_sa[1]['success_rate']:.1f}% success rate)")
    
    def export_results(self):
        """Export results to CSV and detailed Excel with project tabs"""
        if not self.consultant_metrics:
            return
        
        # Create results DataFrame
        results_data = []
        for consultant, metrics in self.consultant_metrics.items():
            results_data.append({
                'Consultant': consultant,
                'Unique_Projects': metrics['unique_projects'],
                'Total_Hours': metrics['total_hours'],
                'Efficiency_Score': metrics['efficiency_score'],
                'Composite_Score': round(metrics['composite_score'], 1),
                'Success_Ratio': metrics['success_ratio'],
                'Projects_Within_Budget': metrics['projects_within_budget'],
                'Projects_Over_Budget': metrics['projects_over_budget'],
                'Projects_On_Hold': metrics['projects_on_hold']
            })
        
        results_df = pd.DataFrame(results_data)
        results_df = results_df.sort_values('Composite_Score', ascending=False)
        
        # Export CSV summary
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'consultant_performance_summary_{timestamp}.csv'
        results_df.to_csv(output_file, index=False)
        print(f"\n[EXPORT] Results exported to: {output_file}")
        
        # Export SA results if available
        if self.sa_metrics:
            self.export_sa_results()
        
        # Export detailed Excel with tabs
        self.export_detailed_excel()
    
    def export_sa_results(self):
        """Export Solution Architect results to CSV"""
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        sa_data = []
        for sa_name, metrics in self.sa_metrics.items():
            sa_data.append({
                'Solution_Architect': sa_name,
                'Total_Projects': metrics['total_projects'],
                'Successful_Projects': metrics['successful_projects'],
                'Failed_Projects': metrics['failed_projects'],
                'Success_Rate': round(metrics['success_rate'], 1),
                'Total_Budgeted_Hours': metrics['total_budgeted_hours'],
                'Total_Actual_Hours': metrics['total_actual_hours'],
                'Composite_Score': round(metrics['composite_score'], 1) if self.show_composite else None
            })
        
        sa_df = pd.DataFrame(sa_data)
        if self.show_composite:
            sa_df = sa_df.sort_values('Composite_Score', ascending=False)
        else:
            if 'Composite_Score' in sa_df.columns:
                sa_df = sa_df.drop('Composite_Score', axis=1)
            sa_df = sa_df.sort_values('Total_Budgeted_Hours', ascending=False)
        
        output_file = f'solution_architect_performance_{timestamp}.csv'
        sa_df.to_csv(output_file, index=False)
        print(f"[EXPORT] Solution Architect results exported to: {output_file}")
        
        # Export detailed SA Excel with tabs
        self.export_detailed_sa_excel()
    
    def export_detailed_excel(self):
        """Export detailed Excel file with consultant project tabs"""
        from datetime import datetime
        from openpyxl.styles import PatternFill
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'consultant_detailed_analysis_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary tab
            results_data = []
            for consultant, metrics in self.consultant_metrics.items():
                results_data.append({
                    'Consultant': consultant,
                    'Unique_Projects': metrics['unique_projects'],
                    'Total_Hours': metrics['total_hours'],
                    'Efficiency_Score': metrics['efficiency_score'],
                    'Composite_Score': round(metrics['composite_score'], 1),
                    'Success_Ratio': metrics['success_ratio'],
                    'Projects_Within_Budget': metrics['projects_within_budget'],
                    'Projects_Over_Budget': metrics['projects_over_budget']
                })
            
            summary_df = pd.DataFrame(results_data)
            if self.show_composite:
                summary_df = summary_df.sort_values('Composite_Score', ascending=False)
            else:
                if 'Composite_Score' in summary_df.columns:
                    summary_df = summary_df.drop('Composite_Score', axis=1)
                summary_df = summary_df.sort_values('Efficiency_Score', ascending=False)
            
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Individual consultant tabs with project details and color coding
            for consultant in sorted(self.consultant_metrics.keys()):
                consultant_projects = self.get_consultant_project_details(consultant)
                if consultant_projects is not None and not consultant_projects.empty:
                    # Add variance calculation column
                    consultant_projects = self.add_variance_column(consultant_projects)
                    
                    # Clean sheet name (Excel tab name restrictions)
                    sheet_name = consultant.replace('/', '_').replace('\\', '_')[:31]
                    consultant_projects.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply color coding
                    self.apply_color_coding(writer.sheets[sheet_name], consultant_projects)
        
        print(f"[EXPORT] Detailed analysis exported to: {output_file}")
    
    def export_detailed_sa_excel(self):
        """Export detailed SA Excel file with individual SA project tabs"""
        from datetime import datetime
        from openpyxl.styles import PatternFill
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'solution_architect_detailed_analysis_{timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary tab
            sa_data = []
            for sa_name, metrics in self.sa_metrics.items():
                sa_data.append({
                    'Solution_Architect': sa_name,
                    'Total_Projects': metrics['total_projects'],
                    'Successful_Projects': metrics['successful_projects'],
                    'Failed_Projects': metrics['failed_projects'],
                    'Success_Rate': round(metrics['success_rate'], 1),
                    'Total_Budgeted_Hours': metrics['total_budgeted_hours'],
                    'Total_Actual_Hours': metrics['total_actual_hours'],
                    'Composite_Score': round(metrics['composite_score'], 1) if self.show_composite else None
                })
            
            summary_df = pd.DataFrame(sa_data)
            if self.show_composite:
                summary_df = summary_df.sort_values('Composite_Score', ascending=False)
            else:
                if 'Composite_Score' in summary_df.columns:
                    summary_df = summary_df.drop('Composite_Score', axis=1)
                summary_df = summary_df.sort_values('Total_Budgeted_Hours', ascending=False)
            
            summary_df.to_excel(writer, sheet_name='SA_Summary', index=False)
            
            # Individual SA tabs with project details
            for sa_name in sorted(self.sa_metrics.keys()):
                sa_projects = self.get_sa_project_details(sa_name)
                if sa_projects is not None and not sa_projects.empty:
                    # Add variance calculation column
                    sa_projects = self.add_variance_column(sa_projects)
                    
                    # Clean sheet name (Excel tab name restrictions)
                    sheet_name = sa_name.replace('/', '_').replace('\\', '_')[:31]
                    sa_projects.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply color coding
                    self.apply_color_coding(writer.sheets[sheet_name], sa_projects)
        
        print(f"[EXPORT] Detailed SA analysis exported to: {output_file}")
    
    def get_consultant_project_details(self, consultant_name):
        """Get detailed project data for a specific consultant"""
        consultant_projects = []
        
        for idx, row in self.data.iterrows():
            resources_col = None
            for col in self.data.columns:
                if 'resource' in str(col).lower() and 'engaged' in str(col).lower():
                    resources_col = col
                    break
            
            if not resources_col:
                continue
                
            resources = row.get(resources_col, "")
            consultants = self.extract_consultants_from_resources(resources)
            
            # Check if this consultant is in this project
            if self.normalize_name(consultant_name) in [self.normalize_name(c) for c in consultants]:
                # Get job number
                job_number = None
                for col in self.data.columns:
                    if 'job' in str(col).lower() and 'number' in str(col).lower():
                        job_number = row.get(col, f"Project_{idx}")
                        break
                
                # Check if excluded
                if not self.is_excluded(consultant_name, job_number):
                    consultant_projects.append(row)
        
        if consultant_projects:
            df = pd.DataFrame(consultant_projects)
            # Remove unnamed columns
            df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
            return df
        return None
    
    def get_sa_project_details(self, sa_name):
        """Get detailed project data for a specific Solution Architect"""
        sa_projects = []
        
        # Find Solution Architect column
        sa_col = None
        for col in self.data.columns:
            if 'solution' in str(col).lower() and 'architect' in str(col).lower():
                sa_col = col
                break
        
        if not sa_col:
            return None
        
        for idx, row in self.data.iterrows():
            sa_field = str(row.get(sa_col, ""))
            if not sa_field or sa_field.lower() in ['nan', 'none', '']:
                continue
            
            # Split by comma and check if this SA is in the list
            sa_names = [self.normalize_name(name.strip()) for name in sa_field.split(',')]
            
            if self.normalize_name(sa_name) in sa_names:
                budgeted_hours = pd.to_numeric(row.get('Budget Hrs', 0), errors='coerce') or 0
                if budgeted_hours > 0:
                    sa_projects.append(row)
        
        if sa_projects:
            df = pd.DataFrame(sa_projects)
            # Remove unnamed columns
            df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
            return df
        return None
    
    def add_variance_column(self, df):
        """Add variance percentage column to project data"""
        actual_col = None
        budget_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if ('actual' in col_lower or 'total' in col_lower or 'posted' in col_lower) and ('hour' in col_lower or 'hrs' in col_lower):
                actual_col = col
            elif 'budget' in col_lower and ('hour' in col_lower or 'hrs' in col_lower):
                budget_col = col
        
        if actual_col and budget_col:
            df['Variance_%'] = df.apply(lambda row: 
                ((pd.to_numeric(row[actual_col], errors='coerce') or 0) - 
                 (pd.to_numeric(row[budget_col], errors='coerce') or 0)) / 
                (pd.to_numeric(row[budget_col], errors='coerce') or 1) * 100 
                if pd.to_numeric(row[budget_col], errors='coerce') and pd.to_numeric(row[budget_col], errors='coerce') > 0 
                else 0, axis=1)
        
        return df
    
    def apply_color_coding(self, worksheet, df):
        """Apply color coding based on variance thresholds"""
        from openpyxl.styles import PatternFill
        
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        
        # Find variance column
        variance_col_idx = None
        for idx, col in enumerate(df.columns):
            if 'Variance_%' in str(col):
                variance_col_idx = idx + 1  # Excel is 1-indexed
                break
        
        if variance_col_idx:
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                variance_cell = worksheet.cell(row=row_idx, column=variance_col_idx)
                try:
                    variance = float(variance_cell.value) if variance_cell.value else 0
                    # Determine color based on variance
                    fill_color = None
                    green_threshold = self.config['thresholds']['green_threshold'] * 100
                    yellow_threshold = self.config['thresholds']['yellow_threshold'] * 100
                    red_threshold = self.config['thresholds']['red_threshold'] * 100
                    
                    if variance < green_threshold:  # Under budget threshold
                        fill_color = green_fill
                    elif variance > red_threshold:  # Over red threshold
                        fill_color = red_fill
                    elif variance > yellow_threshold:  # Over yellow threshold
                        fill_color = yellow_fill
                    
                    # Apply color to entire row if color determined
                    if fill_color:
                        for col_idx in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = fill_color
                except (ValueError, TypeError):
                    pass
    
    def export_excluded_projects(self, excluded_projects):
        """Export projects excluded due to missing budget data"""
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'excluded_projects_{timestamp}.xlsx'
        
        df = pd.DataFrame(excluded_projects)
        # Remove unnamed columns
        df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Excluded_Projects', index=False)
        
        print(f"[EXPORT] Excluded projects exported to: {output_file}")
    
    def check_inconsistent_status_projects(self):
        """
        Check for data quality issues in project status fields.
        
        Identifies two types of problematic projects:
        1. Status inconsistencies: Job Status=Open but Project Status=Closed with old end dates
        2. Missing status data: Project Status is N/A, blank, or undefined
        
        These issues suggest:
        - Projects that should be closed in the system
        - Data entry problems
        - Workflow issues in project management
        
        Prompts user before exporting to avoid unwanted file generation.
        """
        from datetime import datetime, timedelta
        
        inconsistent_projects = []  # Job=Open, Project=Closed, old end date
        na_status_projects = []     # Project Status = N/A or blank
        
        # Get configuration settings
        check_config = self.config.get('inconsistent_project_check', {})
        old_days = check_config.get('old_project_days', 730)  # Default 2 years
        include_na = check_config.get('include_na_status', True)
        
        cutoff_date = datetime.now() - timedelta(days=old_days)
        
        # Find relevant columns
        job_status_col = None
        project_status_col = None
        end_date_col = None
        
        for col in self.data.columns:
            col_lower = str(col).lower()
            if 'job' in col_lower and 'status' in col_lower:
                job_status_col = col
            elif 'project' in col_lower and 'status' in col_lower:
                project_status_col = col
            elif 'end' in col_lower and 'date' in col_lower:
                end_date_col = col
        
        if not all([job_status_col, project_status_col, end_date_col]):
            if self.show_composite:
                print(f"[DEBUG] Could not find required columns for inconsistent status check")
            return
        
        # Count inconsistent projects
        for idx, row in self.data.iterrows():
            job_status = str(row.get(job_status_col, '')).lower()
            project_status = str(row.get(project_status_col, '')).lower()
            end_date_str = row.get(end_date_col, '')
            
            # Check for Job Status = Open AND Project Status = Closed
            if job_status == 'open' and project_status == 'closed' and pd.notna(end_date_str):
                try:
                    # Parse end date
                    end_date = None
                    date_str = str(end_date_str).split()[0]
                    
                    for date_format in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y']:
                        try:
                            end_date = datetime.strptime(date_str, date_format)
                            break
                        except ValueError:
                            continue
                    
                    # Check if end date is older than configured days
                    if end_date and end_date < cutoff_date:
                        inconsistent_projects.append(row)
                        
                except (ValueError, TypeError):
                    continue
            
            # Check for Project Status = N/A if enabled
            elif include_na and project_status in ['n/a', 'na', '']:
                na_status_projects.append(row)
        
        # Auto-export problem projects in Streamlit mode, ask in console mode
        total_issues = len(inconsistent_projects) + len(na_status_projects)
        
        if total_issues > 0:
            days_text = f"{old_days} days" if old_days != 730 else "2 years"
            na_text = f" and {len(na_status_projects)} with N/A status" if na_status_projects else ""
            
            # Auto-export in Streamlit mode, prompt in console mode
            if hasattr(self, 'streamlit_mode') and self.streamlit_mode:
                should_export = True
                print(f"\nFound {len(inconsistent_projects)} projects that are old and need to be reviewed (Job Status=Open, Project Status=Closed, End Date >{days_text}){na_text}.")
                print("Auto-exporting for review in Streamlit mode...")
            else:
                response = input(f"\nFound {len(inconsistent_projects)} projects that are old and need to be reviewed (Job Status=Open, Project Status=Closed, End Date >{days_text}){na_text}.\nDo you want to export these for review? (y/n): ")
                should_export = response.lower() == 'y'
            
            if should_export:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_file = f'inconsistent_status_projects_{timestamp}.xlsx'
                
                # Combine both types of projects
                all_problem_projects = inconsistent_projects + na_status_projects
                
                df = pd.DataFrame(all_problem_projects)
                # Remove unnamed columns
                df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
                
                # Add a column to identify the issue type
                issue_types = ['Old Closed Project'] * len(inconsistent_projects) + ['N/A Status'] * len(na_status_projects)
                df.insert(0, 'Issue_Type', issue_types)
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Problem_Projects', index=False)
                
                print(f"[EXPORT] Problem projects exported to: {output_file}")
            else:
                print(f"[INFO] Skipping export of problem projects")
        else:
            if self.show_composite:
                print(f"[DEBUG] No problem projects found")

def main():
    """
    Main execution function - orchestrates the entire analysis process.
    
    Execution flow:
    1. Parse command line arguments (debug mode)
    2. Initialize analyzer with configuration
    3. Validate input files exist
    4. Load and process data
    5. Run consultant and SA analysis
    6. Generate comprehensive reports
    7. Export detailed Excel files
    8. Run advanced analytics if enabled
    9. Offer file cleanup
    
    Command line options:
    - No args: Normal execution with standard output
    - 'debug' or 'composite': Show composite scores and debug info
    """
    print("Project Financials Consultant Metrics Analyzer")
    print("=" * 60)
    
    # Parse command line arguments
    show_composite = len(sys.argv) > 1 and sys.argv[1].lower() in ['debug', 'composite']
    
    # Initialize analyzer with configuration file
    analyzer = ConsultantMetricsAnalyzer()
    analyzer.show_composite = show_composite
    
    # Validate that main Excel file exists
    if not os.path.exists(analyzer.excel_file):
        print(f"[ERROR] Excel file not found: {analyzer.excel_file}")
        sys.exit(1)
    
    # Execute main analysis pipeline
    analyzer.load_data()           # Load and preprocess data
    analyzer.analyze_consultants() # Analyze consultant performance
    
    # Generate comprehensive reports
    analyzer.generate_report()     # Console report
    analyzer.export_results()      # CSV and Excel exports
    
    # Run advanced analytics if enabled in config
    if analyzer.config.get('advanced_analytics', {}).get('enable_predictive', False):
        analyzer.run_advanced_analytics()
    
    # Run trending analysis if enabled
    if analyzer.config.get('trending_analysis', {}).get('enable_trending', False):
        analyzer.run_trending_analysis()
    
    # Run client analysis if enabled
    if analyzer.config.get('client_analysis', {}).get('enable_client_analysis', False):
        analyzer.run_client_analysis()
    
    # Run DAS+ analysis if enabled
    if analyzer.config.get('das_plus_analysis', {}).get('enable_das_plus', False):
        analyzer.run_das_plus_analysis()
    
    print(f"\n[SUCCESS] Analysis complete!")
    
    # Offer to clean up old report files
    cleanup_old_files()

def cleanup_old_files():
    """
    Clean up old report files to prevent directory clutter.
    
    Manages timestamped files generated by the analyzer:
    - Keeps the most recent file of each type
    - Deletes older versions with timestamps
    - Prompts user before deletion
    
    File types managed:
    - Consultant performance summaries (CSV)
    - Detailed analysis reports (Excel)
    - Solution architect reports (CSV/Excel)
    - Advanced analytics details (Excel)
    - Excluded/inconsistent project reports (Excel)
    """
    import glob
    from pathlib import Path
    
    response = input("\nClean up old CSV and Excel files? Keep only the most recent of each type (y/n): ")
    if response.lower() != 'y':
        return
    
    # Define patterns for all generated report files
    file_patterns = [
        'consultant_performance_summary_*.csv',      # Main consultant CSV reports
        'consultant_detailed_analysis_*.xlsx',       # Detailed consultant Excel reports
        'solution_architect_performance_*.csv',      # SA performance CSV reports
        'solution_architect_detailed_analysis_*.xlsx', # Detailed SA Excel reports
        'excluded_projects_*.xlsx',                  # Projects with missing budget data
        'advanced_analytics_details_*.xlsx',         # Advanced analytics breakdowns
        'inconsistent_status_projects_*.xlsx',       # Data quality issue reports
        'trending_analysis_*.xlsx',                  # Trending analysis reports
        'client_analysis_*.xlsx',                    # Customer performance analysis reports
        'das_plus_analysis_*.xlsx'                   # Combined DAS+ analysis with review projects
    ]
    
    # Process each file type
    for pattern in file_patterns:
        files = glob.glob(pattern)
        if len(files) > 1:
            # Sort by modification time, newest first
            files.sort(key=lambda x: Path(x).stat().st_mtime, reverse=True)
            files_to_delete = files[1:]  # All except the first (newest)
            
            # Delete older files
            for file_path in files_to_delete:
                try:
                    Path(file_path).unlink()
                    print(f"[CLEANUP] Deleted: {file_path}")
                except Exception as e:
                    print(f"[ERROR] Could not delete {file_path}: {e}")
    
    print("[CLEANUP] File cleanup complete!")

# Entry point - run main function when script is executed directly
if __name__ == "__main__":
    main()

# End of consultant_metrics_analyzer_fixed.py
# This comprehensive tool provides detailed analysis of consultant and solution
# architect performance with advanced analytics capabilities.